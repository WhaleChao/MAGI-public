"""Monthly accounting bonus settlement for MAGI/OSC.

The settlement is intentionally idempotent: MAGI can run it on the 24th and
again on later days after the colleague Google Sheet catches up, without
creating duplicate expense rows.
"""

from __future__ import annotations

import calendar
import json
import os
import re
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any


BONUS_SOURCE = "magi_monthly_bonus"
BONUS_CATEGORY = "人事費"
BONUS_SUB_TYPE = "獎金"
LAF_BONUS_LABEL = "法扶消債酬金獎金"
CASE_BONUS_LABEL = "案件獎金"
BONUS_DESC_PREFIX = "[MAGI月結獎金]"
STATUS_LABELS = {
    "waiting_laf_fee": "等待本期法扶消債酬金入帳",
    "ready": "可登載",
    "posted": "已登載",
    "no_surplus_after_laf_bonus": "法扶獎金後無餘額",
    "not_settlement_window": "非結算期間",
}


def _get_osc_helpers():
    from api.osc.utils import _osc_exec

    return _osc_exec


def _as_float(value: Any) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _round_money(value: float) -> float:
    return round(float(value or 0), 2)


def period_for_settlement_month(month: str) -> tuple[date, date, str]:
    m = str(month or "").strip()
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", m)
    if not match:
        raise ValueError("month must be YYYY-MM")
    year, mon = int(match.group(1)), int(match.group(2))
    if mon < 1 or mon > 12:
        raise ValueError("month must be YYYY-MM")
    if mon == 1:
        start = date(year - 1, 12, 26)
    else:
        start = date(year, mon - 1, 26)
    end_day = min(25, calendar.monthrange(year, mon)[1])
    return start, date(year, mon, end_day), f"{year:04d}-{mon:02d}"


def default_settlement_month(today: date | None = None, *, catch_up: bool = False) -> str | None:
    today = today or date.today()
    if today.day >= 24:
        return f"{today.year:04d}-{today.month:02d}"
    if catch_up and today.day <= 7:
        year, mon = today.year, today.month - 1
        if mon == 0:
            year -= 1
            mon = 12
        return f"{year:04d}-{mon:02d}"
    return None


def _settlement_date(month_key: str) -> str:
    year, mon = [int(x) for x in month_key.split("-")]
    day = min(24, calendar.monthrange(year, mon)[1])
    return date(year, mon, day).isoformat()


def _status_label(status: Any) -> str:
    text = str(status or "")
    return STATUS_LABELS.get(text, text)


def _calendar_months_between(start: date, end: date) -> list[str]:
    out: list[str] = []
    year, mon = start.year, start.month
    while (year, mon) <= (end.year, end.month):
        out.append(f"{year:04d}-{mon:02d}")
        mon += 1
        if mon == 13:
            year += 1
            mon = 1
    return out


def _parse_iso_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def ensure_bonus_schema() -> None:
    _osc_exec = _get_osc_helpers()
    _osc_exec(
        """
        CREATE TABLE IF NOT EXISTS accounting_monthly_bonus_runs (
          month_key VARCHAR(7) NOT NULL,
          period_start DATE NOT NULL,
          period_end DATE NOT NULL,
          run_status VARCHAR(40) NOT NULL,
          legal_aid_debt_fee_total DOUBLE NOT NULL DEFAULT 0,
          legal_aid_bonus_amount DOUBLE NOT NULL DEFAULT 0,
          income_total_before_bonus DOUBLE NOT NULL DEFAULT 0,
          expense_total_before_bonus DOUBLE NOT NULL DEFAULT 0,
          balance_after_laf_bonus DOUBLE NOT NULL DEFAULT 0,
          case_bonus_pool DOUBLE NOT NULL DEFAULT 0,
          case_bonus_employee_amount DOUBLE NOT NULL DEFAULT 0,
          final_expense_total DOUBLE NOT NULL DEFAULT 0,
          final_balance DOUBLE NOT NULL DEFAULT 0,
          laf_bonus_transaction_id INT NULL,
          case_bonus_transaction_id INT NULL,
          source_fee_rows_json MEDIUMTEXT NULL,
          import_results_json MEDIUMTEXT NULL,
          xlsx_path TEXT NULL,
          notes TEXT NULL,
          updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          created_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP,
          PRIMARY KEY (month_key)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """,
        fetch="none",
    )
    rows, _ = _osc_exec("SHOW COLUMNS FROM accounting_monthly_bonus_runs", fetch="all")
    fields = {str((row or {}).get("Field") or "") for row in rows or []}
    if "xlsx_path" not in fields:
        _osc_exec(
            "ALTER TABLE accounting_monthly_bonus_runs ADD COLUMN xlsx_path TEXT NULL AFTER import_results_json",
            fetch="none",
        )


def _extract_laf_no(text: str) -> str:
    match = re.search(r"\b\d{7}-[A-Z]-\d{3}\b", str(text or ""))
    return match.group(0) if match else ""


def _extract_client_name_from_fee_description(text: str) -> str:
    raw = str(text or "").strip()
    raw = re.sub(r"^\s*\d{7}-[A-Z]-\d{3}\s*", "", raw)
    raw = raw.split("｜", 1)[0].strip()
    raw = raw.split("-", 1)[0].strip()
    return raw


def _is_debt_case(row: dict[str, Any]) -> bool:
    text = " ".join(
        str(row.get(k) or "")
        for k in ("case_type", "case_category", "case_reason", "case_subject", "folder_path", "folder_name")
    )
    if "消費者債務清理" in text or "消債" in text:
        return True
    return "法律扶助" in text and ("更生" in text or "清算" in text)


def _is_laf_fee_income(row: dict[str, Any]) -> bool:
    if not str(row.get("type") or "").startswith("收入"):
        return False
    if _as_float(row.get("amount")) <= 0:
        return False
    text = " ".join(str(row.get(k) or "") for k in ("category", "sub_type", "description"))
    if BONUS_DESC_PREFIX in text:
        return False
    looks_laf = "法扶" in text or "法律扶助" in text or bool(_extract_laf_no(text))
    looks_fee = any(k in text for k in ("酬金", "領款", "預付", "結案"))
    return looks_laf and looks_fee


def _lookup_case_by_laf_no(laf_no: str) -> dict[str, Any] | None:
    if not laf_no:
        return None
    _osc_exec = _get_osc_helpers()
    row, _ = _osc_exec(
        """
        SELECT id, case_number, client_name, case_type, case_category, case_subject,
               case_reason, folder_path, folder_name, legal_aid_number, laf_case_no
          FROM cases
         WHERE legal_aid_number=%s OR laf_case_no=%s OR application_no=%s
         LIMIT 1
        """,
        (laf_no, laf_no, laf_no),
        fetch="one",
    )
    return row


def _lookup_single_laf_case_by_client_name(client_name: str) -> dict[str, Any] | None:
    name = str(client_name or "").strip()
    if not name:
        return None
    _osc_exec = _get_osc_helpers()
    rows, _ = _osc_exec(
        """
        SELECT id, case_number, client_name, case_type, case_category, case_subject,
               case_reason, folder_path, folder_name, legal_aid_number, laf_case_no
          FROM cases
         WHERE client_name=%s
           AND (
                COALESCE(case_category,'') LIKE '%%法扶%%'
             OR COALESCE(case_category,'') LIKE '%%法律扶助%%'
             OR COALESCE(legal_aid_number,'') <> ''
             OR COALESCE(laf_case_no,'') <> ''
           )
        """,
        (name,),
        fetch="all",
    )
    debt_rows = [dict(r) for r in rows or [] if _is_debt_case(dict(r))]
    if len(debt_rows) == 1:
        return debt_rows[0]
    return None


def query_laf_debt_fee_rows(
    start: date,
    end: date,
    *,
    exclude_transaction_ids: set[int] | None = None,
) -> list[dict[str, Any]]:
    exclude_transaction_ids = exclude_transaction_ids or set()
    _osc_exec = _get_osc_helpers()
    rows, _ = _osc_exec(
        """
        SELECT t.id, t.case_id, t.date, t.type, t.sub_type, t.category, t.description, t.amount,
               c.case_number, c.client_name, c.case_type, c.case_category, c.case_subject,
               c.case_reason, c.folder_path, c.folder_name, c.legal_aid_number, c.laf_case_no
          FROM case_transactions t
          LEFT JOIN cases c ON c.id=t.case_id
         WHERE t.date >= %s
           AND t.date <= %s
           AND t.type LIKE '收入%%'
           AND COALESCE(t.description,'') NOT LIKE %s
           AND (
                COALESCE(t.category,'') LIKE '%%法扶%%'
             OR COALESCE(t.category,'') LIKE '%%法律扶助%%'
             OR COALESCE(t.description,'') LIKE '%%法扶%%'
             OR COALESCE(t.description,'') LIKE '%%法律扶助%%'
             OR COALESCE(t.description,'') LIKE '%%酬金%%'
             OR COALESCE(t.description,'') LIKE '%%領款%%'
             OR COALESCE(t.sub_type,'') LIKE '%%酬金%%'
           )
         ORDER BY t.date ASC, t.id ASC
        """,
        (start.isoformat(), end.isoformat(), f"{BONUS_DESC_PREFIX}%"),
        fetch="all",
    )
    out: list[dict[str, Any]] = []
    seen_ids: set[int] = set()
    for raw in rows or []:
        row = dict(raw)
        if not _is_laf_fee_income(row):
            continue
        if not _is_debt_case(row):
            laf_no = _extract_laf_no(" ".join(str(row.get(k) or "") for k in ("description", "legal_aid_number", "laf_case_no")))
            case = _lookup_case_by_laf_no(laf_no)
            if not case:
                case = _lookup_single_laf_case_by_client_name(
                    _extract_client_name_from_fee_description(str(row.get("description") or ""))
                )
            if case:
                for key, value in case.items():
                    row.setdefault(key, value)
                    if not row.get(key):
                        row[key] = value
        if not _is_debt_case(row):
            continue
        row_id = int(row.get("id") or 0)
        if row_id in exclude_transaction_ids:
            continue
        if row_id in seen_ids:
            continue
        seen_ids.add(row_id)
        row["amount"] = _round_money(_as_float(row.get("amount")))
        out.append(row)
    return out


def _settled_fee_transaction_ids(before_month_key: str) -> set[int]:
    _osc_exec = _get_osc_helpers()
    rows, _ = _osc_exec(
        """
        SELECT month_key, run_status, source_fee_rows_json
          FROM accounting_monthly_bonus_runs
         WHERE month_key < %s
           AND run_status IN ('ready','posted','no_surplus_after_laf_bonus')
           AND legal_aid_bonus_amount > 0
        """,
        (before_month_key,),
        fetch="all",
    )
    out: set[int] = set()
    for row in rows or []:
        try:
            items = json.loads(row.get("source_fee_rows_json") or "[]")
        except Exception:
            items = []
        for item in items if isinstance(items, list) else []:
            tx_id = int((item or {}).get("transaction_id") or 0)
            if tx_id:
                out.add(tx_id)
    return out


def _laf_fee_basis_start(month_key: str, period_start: date, period_end: date) -> date:
    allow_lookback = str(os.environ.get("MAGI_ACCOUNTING_BONUS_ALLOW_LOOKBACK", "")).lower() in {"1", "true", "yes", "on"}
    if not allow_lookback:
        return period_start
    env_start = (os.environ.get("MAGI_ACCOUNTING_BONUS_LAF_FEE_LOOKBACK_START") or "").strip()
    if env_start:
        parsed = _parse_iso_date(env_start)
        if parsed:
            return parsed
    return period_start


def _query_totals_before_bonus(start: date, end: date) -> dict[str, float]:
    _osc_exec = _get_osc_helpers()
    row, _ = _osc_exec(
        """
        SELECT
          COALESCE(SUM(CASE WHEN type LIKE '收入%%' THEN ABS(amount)
                            WHEN amount>=0 AND type NOT LIKE '支出%%' THEN amount ELSE 0 END),0) AS income_total,
          COALESCE(SUM(CASE WHEN type LIKE '支出%%' THEN ABS(amount)
                            WHEN amount<0 THEN ABS(amount) ELSE 0 END),0) AS expense_total
          FROM case_transactions
         WHERE date >= %s
           AND date <= %s
           AND COALESCE(description,'') NOT LIKE %s
        """,
        (start.isoformat(), end.isoformat(), f"{BONUS_DESC_PREFIX}%"),
        fetch="one",
    )
    return {
        "income_total": _round_money(_as_float((row or {}).get("income_total"))),
        "expense_total": _round_money(_as_float((row or {}).get("expense_total"))),
    }


def _bonus_description(month_key: str, label: str) -> str:
    return f"{BONUS_DESC_PREFIX} {month_key} {label}"


def _find_bonus_transaction(month_key: str, label: str) -> int | None:
    _osc_exec = _get_osc_helpers()
    row, _ = _osc_exec(
        """
        SELECT id FROM case_transactions
         WHERE type='支出'
           AND category=%s
           AND sub_type=%s
           AND description=%s
         ORDER BY id DESC
         LIMIT 1
        """,
        (BONUS_CATEGORY, BONUS_SUB_TYPE, _bonus_description(month_key, label)),
        fetch="one",
    )
    return int(row["id"]) if row and row.get("id") else None


def _upsert_bonus_transaction(month_key: str, label: str, amount: float, *, allow_delete_zero: bool) -> int | None:
    _osc_exec = _get_osc_helpers()
    amount = _round_money(amount)
    existing_id = _find_bonus_transaction(month_key, label)
    if amount <= 0:
        if existing_id and allow_delete_zero:
            _osc_exec("DELETE FROM case_transactions WHERE id=%s", (existing_id,), fetch="none")
        return None
    desc = _bonus_description(month_key, label)
    tx_date = _settlement_date(month_key)
    if existing_id:
        _osc_exec(
            """
            UPDATE case_transactions
               SET date=%s, amount=%s, category=%s, sub_type=%s, description=%s
             WHERE id=%s
            """,
            (tx_date, amount, BONUS_CATEGORY, BONUS_SUB_TYPE, desc, existing_id),
            fetch="none",
        )
        return existing_id
    result, _ = _osc_exec(
        """
        INSERT INTO case_transactions (case_id, date, type, sub_type, category, description, amount)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
        """,
        (None, tx_date, "支出", BONUS_SUB_TYPE, BONUS_CATEGORY, desc, amount),
        fetch="none",
    )
    return int((result or {}).get("lastrowid") or 0) or None


def _record_bonus_run(result: dict[str, Any]) -> None:
    _osc_exec = _get_osc_helpers()
    _osc_exec(
        """
        INSERT INTO accounting_monthly_bonus_runs
          (month_key, period_start, period_end, run_status,
           legal_aid_debt_fee_total, legal_aid_bonus_amount,
           income_total_before_bonus, expense_total_before_bonus,
           balance_after_laf_bonus, case_bonus_pool, case_bonus_employee_amount,
           final_expense_total, final_balance,
           laf_bonus_transaction_id, case_bonus_transaction_id,
           source_fee_rows_json, import_results_json, xlsx_path, notes)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
           period_start=VALUES(period_start),
           period_end=VALUES(period_end),
           run_status=VALUES(run_status),
           legal_aid_debt_fee_total=VALUES(legal_aid_debt_fee_total),
           legal_aid_bonus_amount=VALUES(legal_aid_bonus_amount),
           income_total_before_bonus=VALUES(income_total_before_bonus),
           expense_total_before_bonus=VALUES(expense_total_before_bonus),
           balance_after_laf_bonus=VALUES(balance_after_laf_bonus),
           case_bonus_pool=VALUES(case_bonus_pool),
           case_bonus_employee_amount=VALUES(case_bonus_employee_amount),
           final_expense_total=VALUES(final_expense_total),
           final_balance=VALUES(final_balance),
           laf_bonus_transaction_id=VALUES(laf_bonus_transaction_id),
           case_bonus_transaction_id=VALUES(case_bonus_transaction_id),
           source_fee_rows_json=VALUES(source_fee_rows_json),
           import_results_json=VALUES(import_results_json),
           xlsx_path=VALUES(xlsx_path),
           notes=VALUES(notes)
        """,
        (
            result["month"],
            result["period_start"],
            result["period_end"],
            result["status"],
            result["legal_aid_debt_fee_total"],
            result["legal_aid_bonus_amount"],
            result["income_total_before_bonus"],
            result["expense_total_before_bonus"],
            result["balance_after_laf_bonus"],
            result["case_bonus_pool"],
            result["case_bonus_employee_amount"],
            result["final_expense_total"],
            result["final_balance"],
            result.get("laf_bonus_transaction_id"),
            result.get("case_bonus_transaction_id"),
            json.dumps(result.get("source_fee_rows") or [], ensure_ascii=False),
            json.dumps(result.get("import_results") or [], ensure_ascii=False),
            result.get("xlsx_path") or None,
            result.get("notes") or "",
        ),
        fetch="none",
    )


def record_monthly_bonus_xlsx_path(month_key: str, xlsx_path: str | Path) -> None:
    ensure_bonus_schema()
    _osc_exec = _get_osc_helpers()
    _osc_exec(
        "UPDATE accounting_monthly_bonus_runs SET xlsx_path=%s WHERE month_key=%s",
        (str(xlsx_path), month_key),
        fetch="none",
    )


def refresh_accounting_import_for_period(
    start: date,
    end: date,
    *,
    commit: bool,
    account_hint: str | None = None,
) -> list[dict[str, Any]]:
    from api.osc.accounting_sheet_import import DEFAULT_ACCOUNT_HINT, DEFAULT_GID, DEFAULT_SPREADSHEET_ID, run_import

    results: list[dict[str, Any]] = []
    for month in _calendar_months_between(start, end):
        try:
            results.append(
                run_import(
                    month=month,
                    dry_run=not commit,
                    spreadsheet_id=DEFAULT_SPREADSHEET_ID,
                    gid=DEFAULT_GID,
                    interactive=False,
                    account_hint=account_hint or DEFAULT_ACCOUNT_HINT,
                )
            )
        except Exception as exc:
            results.append({"ok": False, "month": month, "error": type(exc).__name__, "message": str(exc)})
    return results


def calculate_monthly_bonus(
    *,
    month: str | None = None,
    today: date | None = None,
    commit: bool = False,
    refresh_import: bool = True,
    catch_up: bool = False,
    account_hint: str | None = None,
) -> dict[str, Any]:
    ensure_bonus_schema()
    month_key = month or default_settlement_month(today, catch_up=catch_up)
    if not month_key:
        return {
            "ok": True,
            "skipped": True,
            "status": "not_settlement_window",
            "message": "今天不是月結獎金結算或月初重算期間。",
        }
    start, end, month_key = period_for_settlement_month(month_key)
    import_results = refresh_accounting_import_for_period(start, end, commit=commit, account_hint=account_hint) if refresh_import else []

    settled_fee_ids = _settled_fee_transaction_ids(month_key)
    fee_basis_start = _laf_fee_basis_start(month_key, start, end)
    fee_rows = query_laf_debt_fee_rows(fee_basis_start, end, exclude_transaction_ids=settled_fee_ids)
    fee_total = _round_money(sum(_as_float(r.get("amount")) for r in fee_rows))
    legal_aid_bonus = _round_money(fee_total * float(os.environ.get("MAGI_ACCOUNTING_LAF_BONUS_RATE", "0.5") or "0.5"))
    totals = _query_totals_before_bonus(start, end)
    period_income_before = totals["income_total"]
    expense_before = totals["expense_total"]
    prior_period_fee_income = _round_money(
        sum(
            _as_float(r.get("amount"))
            for r in fee_rows
            if (tx_date := _parse_iso_date(r.get("date"))) and tx_date < start
        )
    )
    income_before = _round_money(period_income_before + prior_period_fee_income)
    require_laf_fee = str(os.environ.get("MAGI_ACCOUNTING_BONUS_REQUIRE_LAF_FEE", "1")).lower() in {"1", "true", "yes", "on"}

    balance_after_laf = _round_money(income_before - expense_before - legal_aid_bonus)
    case_pool_rate = float(os.environ.get("MAGI_ACCOUNTING_CASE_BONUS_POOL_RATE", "0.5") or "0.5")
    case_employee_rate = float(os.environ.get("MAGI_ACCOUNTING_CASE_BONUS_EMPLOYEE_RATE", "0.5") or "0.5")
    case_bonus_pool = _round_money(max(balance_after_laf, 0) * case_pool_rate)
    case_bonus_employee = _round_money(case_bonus_pool * case_employee_rate)
    final_expense = _round_money(expense_before + legal_aid_bonus + case_bonus_employee)
    final_balance = _round_money(income_before - final_expense)

    status = "ready"
    notes = ""
    if require_laf_fee and fee_total <= 0:
        status = "waiting_laf_fee"
        notes = "本次結算尚未在帳務收入中找到尚未計獎的法扶消債酬金；MAGI 會在 24 日後重新匯入同事帳務並重算。"
        legal_aid_bonus = 0.0
        case_bonus_pool = 0.0
        case_bonus_employee = 0.0
        final_expense = expense_before
        final_balance = _round_money(income_before - expense_before)
        balance_after_laf = final_balance
    elif balance_after_laf <= 0:
        status = "no_surplus_after_laf_bonus"
        notes = "法扶酬金獎金計入後本期沒有正餘額，因此不登載案件獎金。"

    laf_tx_id = _find_bonus_transaction(month_key, LAF_BONUS_LABEL)
    case_tx_id = _find_bonus_transaction(month_key, CASE_BONUS_LABEL)
    if commit and status != "waiting_laf_fee":
        laf_tx_id = _upsert_bonus_transaction(month_key, LAF_BONUS_LABEL, legal_aid_bonus, allow_delete_zero=False)
        case_tx_id = _upsert_bonus_transaction(month_key, CASE_BONUS_LABEL, case_bonus_employee, allow_delete_zero=True)
        status = "posted" if legal_aid_bonus > 0 or case_bonus_employee > 0 else status
    elif commit:
        case_tx_id = _upsert_bonus_transaction(month_key, CASE_BONUS_LABEL, 0, allow_delete_zero=True)

    result = {
        "ok": True,
        "dry_run": not commit,
        "month": month_key,
        "period_start": start.isoformat(),
        "period_end": end.isoformat(),
        "fee_basis_start": fee_basis_start.isoformat(),
        "fee_basis_end": end.isoformat(),
        "settlement_date": _settlement_date(month_key),
        "status": status,
        "notes": notes,
        "legal_aid_debt_fee_total": fee_total,
        "legal_aid_bonus_amount": legal_aid_bonus,
        "legal_aid_bonus_rate": float(os.environ.get("MAGI_ACCOUNTING_LAF_BONUS_RATE", "0.5") or "0.5"),
        "period_income_total_before_bonus": period_income_before,
        "unsettled_laf_fee_income_before_period": prior_period_fee_income,
        "income_total_before_bonus": income_before,
        "expense_total_before_bonus": expense_before,
        "balance_after_laf_bonus": balance_after_laf,
        "case_bonus_pool_rate": case_pool_rate,
        "case_bonus_employee_rate": case_employee_rate,
        "case_bonus_pool": case_bonus_pool,
        "case_bonus_employee_amount": case_bonus_employee,
        "final_expense_total": final_expense,
        "final_balance": final_balance,
        "laf_bonus_transaction_id": laf_tx_id,
        "case_bonus_transaction_id": case_tx_id,
        "source_fee_rows": [
            {
                "transaction_id": r.get("id"),
                "date": r.get("date"),
                "amount": r.get("amount"),
                "case_number": r.get("case_number"),
                "client_name": r.get("client_name"),
                "laf_case_no": r.get("laf_case_no") or r.get("legal_aid_number") or _extract_laf_no(str(r.get("description") or "")),
                "case_reason": r.get("case_reason"),
                "description": r.get("description"),
            }
            for r in fee_rows
        ],
        "import_results": import_results,
    }
    if commit:
        _record_bonus_run(result)
    return result


def export_monthly_bonus_xlsx(result: dict[str, Any], output_path: str | Path | None = None) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError(f"openpyxl unavailable: {exc}") from exc

    if output_path is None:
        out_dir = Path(os.environ.get("MAGI_EXPORT_DIR") or Path.cwd() / "static" / "exports")
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / f"magi_accounting_bonus_{result.get('month')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "月結獎金"
    rows = [
        ("結算月份", result.get("month")),
        ("期間", f"{result.get('period_start')} ~ {result.get('period_end')}"),
        ("法扶消債酬金計算範圍", f"{result.get('fee_basis_start') or result.get('period_start')} ~ {result.get('fee_basis_end') or result.get('period_end')}（排除先前已計獎交易）"),
        ("狀態", _status_label(result.get("status"))),
        ("法扶消債酬金收入", result.get("legal_aid_debt_fee_total")),
        ("法扶酬金獎金（收入的一半）", result.get("legal_aid_bonus_amount")),
        ("本期帳務收入", result.get("period_income_total_before_bonus", result.get("income_total_before_bonus"))),
        ("本期前尚未計獎法扶消債酬金", result.get("unsettled_laf_fee_income_before_period", 0)),
        ("獎金前收入合計", result.get("income_total_before_bonus")),
        ("獎金前支出合計", result.get("expense_total_before_bonus")),
        ("法扶獎金後餘額", result.get("balance_after_laf_bonus")),
        ("案件獎金池（餘額的一半）", result.get("case_bonus_pool")),
        ("員工案件獎金（獎金池的一半）", result.get("case_bonus_employee_amount")),
        ("本月支出總額（含本次獎金）", result.get("final_expense_total")),
        ("本月結餘", result.get("final_balance")),
        ("備註", result.get("notes") or ""),
    ]
    ws.append(["項目", "內容"])
    for row in rows:
        ws.append(list(row))

    fee_ws = wb.create_sheet("法扶消債酬金明細")
    fee_ws.append(["交易ID", "日期", "本所案號", "當事人", "法扶案號", "案由", "說明", "金額"])
    for item in result.get("source_fee_rows") or []:
        fee_ws.append([
            item.get("transaction_id"),
            item.get("date"),
            item.get("case_number"),
            item.get("client_name"),
            item.get("laf_case_no"),
            item.get("case_reason"),
            item.get("description"),
            item.get("amount"),
        ])

    import_ws = wb.create_sheet("帳務匯入紀錄")
    import_ws.append(["月份", "狀態", "可匯入", "已匯入過", "DB已有", "固定支出跳過", "錯誤"])
    for item in result.get("import_results") or []:
        import_ws.append([
            item.get("month"),
            "成功" if item.get("ok") else "失敗",
            item.get("importable_count", 0),
            item.get("duplicate_count", 0),
            item.get("existing_count", 0),
            item.get("fixed_expense_skip_count", 0),
            item.get("message") or item.get("error") or "",
        ])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            max_len = 12
            for cell in sheet[letter]:
                max_len = max(max_len, min(45, len(str(cell.value or "")) + 2))
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0;(#,##0);-'
            sheet.column_dimensions[letter].width = max_len
    wb.save(output_path)
    return str(output_path)


def write_temp_xlsx(result: dict[str, Any]) -> str:
    tmp = tempfile.NamedTemporaryFile(prefix="magi_accounting_bonus_", suffix=".xlsx", delete=False)
    tmp.close()
    return export_monthly_bonus_xlsx(result, tmp.name)
