"""
OSC Accounting Blueprint
========================
Handles /api/osc/accounting/* routes: transactions, summary, defaults, recurring.
Migrated from server.py to reduce monolith size.
"""

import logging
from datetime import date, datetime
import os
import tempfile
from pathlib import Path

from flask import Blueprint, request, jsonify, send_file, after_this_request, current_app
from flask_login import login_required

from api.osc.accounting_summary import load_accounting_summary

osc_accounting_bp = Blueprint("osc_accounting", __name__)


def _require_accounting_operator():
    if current_app.config.get("LOGIN_DISABLED"):
        return None
    from api.authz import check_authorization

    allowed, reason = check_authorization("operator")
    if allowed:
        return None
    return jsonify({"ok": False, "error": "forbidden", "reason": reason}), 403


def _get_osc_helpers():
    """Lazy import OSC helpers from server.py to avoid circular imports."""
    from api.osc.utils import _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int
    return _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int


# ── Transactions ─────────────────────────────────────────────────────

def _accounting_transaction_filters(args, *, default_limit: int = 300, max_limit: int = 1000):
    q = (args.get("q") or "").strip()
    case_id = (args.get("case_number") or args.get("case_id") or "").strip()
    limit = max(1, min(max_limit, int(args.get("limit") or str(default_limit))))
    start_date = (args.get("start_date") or "").strip()
    end_date = (args.get("end_date") or "").strip()
    where = []
    params = []
    if case_id:
        where.append("(t.case_id=%s OR t.case_id IN (SELECT id FROM cases WHERE case_number=%s))")
        params.extend([case_id, case_id])
    if start_date:
        where.append("t.date >= %s")
        params.append(start_date)
    if end_date:
        where.append("t.date <= %s")
        params.append(end_date)
    if q:
        like = f"%{q}%"
        where.append("(t.case_id LIKE %s OR t.type LIKE %s OR t.sub_type LIKE %s OR t.category LIKE %s OR t.description LIKE %s)")
        params.extend([like, like, like, like, like])
    return where, params, limit, {"q": q, "case_id": case_id, "start_date": start_date, "end_date": end_date}


def _accounting_transactions_sql(where, *, order: str = "DESC") -> str:
    sql = """
        SELECT t.id, t.case_id, c.case_number, t.date, t.type, t.sub_type, t.category, t.description, t.amount
        FROM case_transactions t
        LEFT JOIN cases c ON c.id = t.case_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    direction = "ASC" if str(order).upper() == "ASC" else "DESC"
    sql += f" ORDER BY t.date {direction}, t.id {direction} LIMIT %s"
    return sql


def _as_amount(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _iso_date_text(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value or "")


@osc_accounting_bp.route("/api/osc/accounting/transactions", methods=["GET", "POST"])
@login_required
def osc_accounting_transactions_api():
    if request.method == "POST":
        auth_error = _require_accounting_operator()
        if auth_error:
            return auth_error
    _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int = _get_osc_helpers()
    if request.method == "GET":
        where, params, limit, _filters = _accounting_transaction_filters(request.args, default_limit=300, max_limit=1000)
        sql = _accounting_transactions_sql(where, order="DESC")
        params.append(limit)
        rows, _ = _osc_exec(sql, tuple(params), fetch="all")
        return jsonify({"ok": True, "items": rows})

    payload = request.get_json() or {}
    raw_case_id = str(payload.get("case_id") or payload.get("case_number") or "").strip()
    case_id = _osc_resolve_case_id(raw_case_id) if raw_case_id else None
    tx_date = str(payload.get("date") or "").strip() or str(date.today())
    if raw_case_id and not case_id:
        return jsonify({"ok": False, "error": "case not found", "message": "找不到這個案件編號，若是共同收入或支出請留空。"}), 400
    try:
        amount = float(payload.get("amount") or 0)
    except Exception:
        return jsonify({"ok": False, "error": "amount invalid"}), 400
    cols = ["case_id", "date", "type", "sub_type", "category", "description", "amount"]
    vals = [
        case_id,
        tx_date,
        str(payload.get("type") or "").strip() or None,
        str(payload.get("sub_type") or "").strip() or None,
        str(payload.get("category") or "").strip() or None,
        str(payload.get("description") or "").strip() or None,
        amount,
    ]
    result, _ = _osc_exec(
        f"INSERT INTO case_transactions ({','.join(cols)}) VALUES ({','.join(['%s'] * len(cols))})",
        tuple(vals),
        fetch="none",
    )
    return jsonify({"ok": True, "result": result})


@osc_accounting_bp.route("/api/osc/accounting/transactions/xlsx", methods=["GET"])
@login_required
def osc_accounting_transactions_xlsx_api():
    _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int = _get_osc_helpers()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        return jsonify({"ok": False, "error": "openpyxl_unavailable", "message": str(exc)}), 500

    try:
        where, params, limit, filters = _accounting_transaction_filters(request.args, default_limit=5000, max_limit=20000)
    except Exception as exc:
        return jsonify({"ok": False, "error": "invalid_filter", "message": str(exc)}), 400
    rows, _ = _osc_exec(_accounting_transactions_sql(where, order="ASC"), tuple([*params, limit]), fetch="all")
    rows = rows or []

    income_total = 0.0
    expense_total = 0.0
    for row in rows:
        amount = abs(_as_amount(row.get("amount")))
        tx_type = str(row.get("type") or "")
        if "支出" in tx_type:
            expense_total += amount
        elif "收入" in tx_type:
            income_total += amount
        elif _as_amount(row.get("amount")) < 0:
            expense_total += amount
        else:
            income_total += amount
    net_total = income_total - expense_total

    wb = Workbook()
    ws = wb.active
    ws.title = "帳務明細"
    headers = ["ID", "日期", "本所案號", "DB案件ID", "收入/支出", "細項", "分類", "說明", "金額"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row.get("id"),
            _iso_date_text(row.get("date")),
            row.get("case_number") or "",
            row.get("case_id") or "",
            row.get("type") or "",
            row.get("sub_type") or "",
            row.get("category") or "",
            row.get("description") or "",
            _as_amount(row.get("amount")),
        ])

    summary = wb.create_sheet("摘要")
    summary.append(["項目", "內容"])
    summary_rows = [
        ("匯出時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("起日", filters.get("start_date") or "未指定"),
        ("迄日", filters.get("end_date") or "未指定"),
        ("案件編號", filters.get("case_id") or "全部"),
        ("關鍵字", filters.get("q") or "無"),
        ("筆數", len(rows)),
        ("收入合計", income_total),
        ("支出合計", expense_total),
        ("淨額", net_total),
        ("匯出上限", limit),
    ]
    for item in summary_rows:
        summary.append(list(item))

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            width = 12
            for cell in sheet[letter]:
                width = max(width, min(48, len(str(cell.value or "")) + 2))
                if isinstance(cell.value, (int, float)) and sheet.title != "摘要":
                    cell.number_format = '#,##0;(#,##0);-'
                elif isinstance(cell.value, (int, float)) and sheet.title == "摘要":
                    cell.number_format = '#,##0.##;(#,##0.##);-'
            sheet.column_dimensions[letter].width = width

    tmp = tempfile.NamedTemporaryFile(prefix="magi_accounting_transactions_", suffix=".xlsx", delete=False)
    tmp.close()
    path = Path(tmp.name)
    wb.save(path)

    @after_this_request
    def _cleanup(response):
        try:
            os.unlink(path)
        except Exception:
            logging.getLogger(__name__).warning("nonfatal exception was ignored at %s:%s", __name__, 210, exc_info=True)
        return response

    start = filters.get("start_date") or "all"
    end = filters.get("end_date") or "all"
    filename = f"MAGI帳務明細_{start}_{end}.xlsx"
    return send_file(path, as_attachment=True, download_name=filename)


@osc_accounting_bp.route("/api/osc/accounting/transactions/<int:row_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def osc_accounting_transaction_detail_api(row_id):
    if request.method != "GET":
        auth_error = _require_accounting_operator()
        if auth_error:
            return auth_error
    _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int = _get_osc_helpers()
    if request.method == "GET":
        row, _ = _osc_exec("SELECT * FROM case_transactions WHERE id=%s", (row_id,), fetch="one")
        if not row:
            return jsonify({"ok": False, "error": "not found"}), 404
        return jsonify({"ok": True, "item": row})
    if request.method == "DELETE":
        result, _ = _osc_exec("DELETE FROM case_transactions WHERE id=%s", (row_id,), fetch="none")
        return jsonify({"ok": True, "result": result})
    payload = request.get_json() or {}
    allowed = ["case_id", "date", "type", "sub_type", "category", "description", "amount"]
    sets = []
    vals = []
    for k in allowed:
        if k not in payload:
            continue
        sets.append(f"{k}=%s")
        if k == "amount":
            try:
                vals.append(float(payload.get(k) or 0))
            except Exception:
                return jsonify({"ok": False, "error": "amount invalid"}), 400
        else:
            v = (payload.get(k) or "").strip() or None
            if k == "case_id" and v:
                resolved = _osc_resolve_case_id(v)
                if not resolved:
                    return jsonify({"ok": False, "error": "case not found", "message": "找不到這個案件編號，若是共同收入或支出請留空。"}), 400
                v = resolved
            vals.append(v)
    if not sets:
        return jsonify({"ok": False, "error": "no fields"}), 400
    vals.append(row_id)
    result, _ = _osc_exec(f"UPDATE case_transactions SET {','.join(sets)} WHERE id=%s", tuple(vals), fetch="none")
    return jsonify({"ok": True, "result": result})


# ── Summary ──────────────────────────────────────────────────────────

@osc_accounting_bp.route("/api/osc/accounting/summary", methods=["GET"])
@login_required
def osc_accounting_summary_api():
    _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int = _get_osc_helpers()
    case_id = (request.args.get("case_number") or request.args.get("case_id") or "").strip()
    start_date = (request.args.get("start_date") or "").strip()
    end_date = (request.args.get("end_date") or "").strip()
    summary = load_accounting_summary(
        _osc_exec,
        case_id=case_id,
        start_date=start_date,
        end_date=end_date,
    )
    return jsonify({"ok": True, **summary})


@osc_accounting_bp.route("/api/osc/accounting/import/google-sheet", methods=["GET", "POST"])
@login_required
def osc_accounting_google_sheet_import_api():
    payload = request.get_json(silent=True) or {}
    commit = bool(payload.get("commit")) if request.method == "POST" else False
    auth = bool(payload.get("auth")) if request.method == "POST" else False
    if request.method == "POST" and (commit or auth):
        auth_error = _require_accounting_operator()
        if auth_error:
            return auth_error

    from api.osc.accounting_sheet_import import (
        DEFAULT_ACCOUNT_HINT,
        DEFAULT_GID,
        DEFAULT_SPREADSHEET_ID,
        SheetsAuthorizationRequired,
        run_import,
    )

    month = (payload.get("month") or request.args.get("month") or "").strip() or None
    try:
        result = run_import(
            month=month,
            dry_run=not commit,
            spreadsheet_id=(payload.get("spreadsheet_id") or request.args.get("spreadsheet_id") or DEFAULT_SPREADSHEET_ID),
            gid=int(payload.get("gid") or request.args.get("gid") or DEFAULT_GID),
            interactive=auth,
            account_hint=(payload.get("account_hint") or request.args.get("account_hint") or DEFAULT_ACCOUNT_HINT),
        )
    except SheetsAuthorizationRequired as exc:
        return jsonify({"ok": False, "error": "auth_required", "message": str(exc)}), 428
    except Exception as exc:
        return jsonify({"ok": False, "error": type(exc).__name__, "message": str(exc)}), 500
    return jsonify(result)


@osc_accounting_bp.route("/api/osc/accounting/monthly-bonus", methods=["GET", "POST"])
@login_required
def osc_accounting_monthly_bonus_api():
    payload = request.get_json(silent=True) or {}
    month = (payload.get("month") or request.args.get("month") or "").strip() or None
    commit = bool(payload.get("commit")) if request.method == "POST" else False
    if commit:
        auth_error = _require_accounting_operator()
        if auth_error:
            return auth_error

    from api.osc.accounting_bonus import (
        calculate_monthly_bonus,
        export_monthly_bonus_xlsx,
        record_monthly_bonus_xlsx_path,
    )
    from api.osc.accounting_sheet_import import DEFAULT_ACCOUNT_HINT

    refresh_raw = payload.get("refresh_import") if "refresh_import" in payload else request.args.get("refresh_import", "1")
    refresh_import = str(refresh_raw).strip().lower() not in {"0", "false", "no", "off"}
    account_hint = (payload.get("account_hint") or request.args.get("account_hint") or DEFAULT_ACCOUNT_HINT).strip()
    try:
        result = calculate_monthly_bonus(
            month=month,
            commit=commit,
            refresh_import=refresh_import,
            catch_up=True,
            account_hint=account_hint,
        )
        if commit and result.get("ok") and not result.get("skipped"):
            result["xlsx_path"] = export_monthly_bonus_xlsx(result)
            record_monthly_bonus_xlsx_path(str(result.get("month") or ""), result["xlsx_path"])
    except Exception as exc:
        return jsonify({"ok": False, "error": type(exc).__name__, "message": str(exc)}), 500
    return jsonify(result)


@osc_accounting_bp.route("/api/osc/accounting/monthly-bonus/xlsx", methods=["GET"])
@login_required
def osc_accounting_monthly_bonus_xlsx_api():
    from api.osc.accounting_bonus import calculate_monthly_bonus, write_temp_xlsx
    from api.osc.accounting_sheet_import import DEFAULT_ACCOUNT_HINT

    month = (request.args.get("month") or "").strip() or None
    refresh_import = str(request.args.get("refresh_import") or "0").strip().lower() in {"1", "true", "yes", "on"}
    account_hint = (request.args.get("account_hint") or DEFAULT_ACCOUNT_HINT).strip()
    try:
        result = calculate_monthly_bonus(
            month=month,
            commit=False,
            refresh_import=refresh_import,
            catch_up=True,
            account_hint=account_hint,
        )
        path = write_temp_xlsx(result)
    except Exception as exc:
        return jsonify({"ok": False, "error": type(exc).__name__, "message": str(exc)}), 500
    filename = f"MAGI帳務月結獎金_{result.get('month') or '未指定'}.xlsx"
    return send_file(path, as_attachment=True, download_name=filename)


# ── Expense Defaults ─────────────────────────────────────────────────

@osc_accounting_bp.route("/api/osc/accounting/defaults", methods=["GET", "POST"])
@login_required
def osc_accounting_defaults_api():
    if request.method == "POST":
        auth_error = _require_accounting_operator()
        if auth_error:
            return auth_error
    _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int = _get_osc_helpers()
    if request.method == "GET":
        q = (request.args.get("q") or "").strip()
        limit = max(1, min(1000, int(request.args.get("limit") or "300")))
        sql = "SELECT id, category, default_description, default_amount FROM expense_defaults WHERE 1=1 "
        params = []
        if q:
            like = f"%{q}%"
            sql += "AND (category LIKE %s OR default_description LIKE %s) "
            params.extend([like, like])
        sql += "ORDER BY category ASC, id DESC LIMIT %s"
        params.append(limit)
        rows, _ = _osc_exec(sql, tuple(params), fetch="all")
        return jsonify({"ok": True, "items": rows})
    payload = request.get_json() or {}
    category = (payload.get("category") or "").strip()
    if not category:
        return jsonify({"ok": False, "error": "category required"}), 400
    try:
        amt = float(payload.get("default_amount") or 0)
    except Exception:
        return jsonify({"ok": False, "error": "default_amount invalid"}), 400
    result, _ = _osc_exec(
        "INSERT INTO expense_defaults (category, default_description, default_amount) VALUES (%s,%s,%s)",
        (category, (payload.get("default_description") or "").strip() or None, amt),
        fetch="none",
    )
    return jsonify({"ok": True, "result": result})


@osc_accounting_bp.route("/api/osc/accounting/defaults/<int:row_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def osc_accounting_default_detail_api(row_id):
    if request.method != "GET":
        auth_error = _require_accounting_operator()
        if auth_error:
            return auth_error
    _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int = _get_osc_helpers()
    if request.method == "GET":
        row, _ = _osc_exec("SELECT * FROM expense_defaults WHERE id=%s", (row_id,), fetch="one")
        if not row:
            return jsonify({"ok": False, "error": "not found"}), 404
        return jsonify({"ok": True, "item": row})
    if request.method == "DELETE":
        result, _ = _osc_exec("DELETE FROM expense_defaults WHERE id=%s", (row_id,), fetch="none")
        return jsonify({"ok": True, "result": result})
    payload = request.get_json() or {}
    sets, vals = [], []
    for k in ["category", "default_description", "default_amount"]:
        if k not in payload:
            continue
        sets.append(f"{k}=%s")
        if k == "default_amount":
            try:
                vals.append(float(payload.get(k) or 0))
            except Exception:
                return jsonify({"ok": False, "error": "default_amount invalid"}), 400
        else:
            vals.append((payload.get(k) or "").strip() or None)
    if not sets:
        return jsonify({"ok": False, "error": "no fields"}), 400
    vals.append(row_id)
    result, _ = _osc_exec(f"UPDATE expense_defaults SET {','.join(sets)} WHERE id=%s", tuple(vals), fetch="none")
    return jsonify({"ok": True, "result": result})


# ── Recurring Expenses ───────────────────────────────────────────────

@osc_accounting_bp.route("/api/osc/accounting/recurring", methods=["GET", "POST"])
@login_required
def osc_accounting_recurring_api():
    if request.method == "POST":
        auth_error = _require_accounting_operator()
        if auth_error:
            return auth_error
    _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int = _get_osc_helpers()
    if request.method == "GET":
        q = (request.args.get("q") or "").strip()
        only_active = str(request.args.get("only_active") or "0").strip().lower() in {"1", "true", "yes", "on"}
        limit = max(1, min(1000, int(request.args.get("limit") or "300")))
        sql = (
            "SELECT id, category, sub_type, description, amount, day_of_month, start_date, end_date, is_active, last_generated_month, created_date "
            "FROM recurring_expenses WHERE 1=1 "
        )
        params = []
        if only_active:
            sql += "AND is_active=1 "
        if q:
            like = f"%{q}%"
            sql += "AND (category LIKE %s OR sub_type LIKE %s OR description LIKE %s) "
            params.extend([like, like, like])
        sql += "ORDER BY is_active DESC, category ASC, id DESC LIMIT %s"
        params.append(limit)
        rows, _ = _osc_exec(sql, tuple(params), fetch="all")
        return jsonify({"ok": True, "items": rows})

    payload = request.get_json() or {}
    category = (payload.get("category") or "").strip()
    if not category:
        return jsonify({"ok": False, "error": "category required"}), 400
    try:
        amount = float(payload.get("amount") or 0)
    except Exception:
        return jsonify({"ok": False, "error": "amount invalid"}), 400
    day_of_month = _osc_safe_int(payload.get("day_of_month"), 1)
    if day_of_month < 1 or day_of_month > 31:
        return jsonify({"ok": False, "error": "day_of_month invalid"}), 400
    is_active = 1 if str(payload.get("is_active") or "").strip().lower() in {"1", "true", "yes", "on"} else 0
    result, _ = _osc_exec(
        "INSERT INTO recurring_expenses (category, sub_type, description, amount, day_of_month, start_date, end_date, is_active, last_generated_month) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (
            category,
            (payload.get("sub_type") or "").strip() or None,
            (payload.get("description") or "").strip() or None,
            amount,
            day_of_month,
            (payload.get("start_date") or "").strip() or None,
            (payload.get("end_date") or "").strip() or None,
            is_active,
            (payload.get("last_generated_month") or "").strip() or None,
        ),
        fetch="none",
    )
    return jsonify({"ok": True, "result": result})


@osc_accounting_bp.route("/api/osc/accounting/recurring/<int:row_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def osc_accounting_recurring_detail_api(row_id):
    if request.method != "GET":
        auth_error = _require_accounting_operator()
        if auth_error:
            return auth_error
    _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int = _get_osc_helpers()
    if request.method == "GET":
        row, _ = _osc_exec("SELECT * FROM recurring_expenses WHERE id=%s", (row_id,), fetch="one")
        if not row:
            return jsonify({"ok": False, "error": "not found"}), 404
        return jsonify({"ok": True, "item": row})
    if request.method == "DELETE":
        result, _ = _osc_exec("DELETE FROM recurring_expenses WHERE id=%s", (row_id,), fetch="none")
        return jsonify({"ok": True, "result": result})
    payload = request.get_json() or {}
    allowed = ["category", "sub_type", "description", "amount", "day_of_month", "start_date", "end_date", "is_active", "last_generated_month"]
    sets, vals = [], []
    for k in allowed:
        if k not in payload:
            continue
        sets.append(f"{k}=%s")
        if k in {"amount"}:
            try:
                vals.append(float(payload.get(k) or 0))
            except Exception:
                return jsonify({"ok": False, "error": f"{k} invalid"}), 400
        elif k in {"day_of_month", "is_active"}:
            vals.append(_osc_safe_int(payload.get(k), 0))
        else:
            vals.append((payload.get(k) or "").strip() or None)
    if not sets:
        return jsonify({"ok": False, "error": "no fields"}), 400
    vals.append(row_id)
    result, _ = _osc_exec(f"UPDATE recurring_expenses SET {','.join(sets)} WHERE id=%s", tuple(vals), fetch="none")
    return jsonify({"ok": True, "result": result})


@osc_accounting_bp.route("/api/osc/accounting/recurring/<int:row_id>/sync-generated", methods=["POST"])
@login_required
def osc_accounting_recurring_sync_generated_api(row_id):
    auth_error = _require_accounting_operator()
    if auth_error:
        return auth_error
    _osc_exec, _osc_text, _osc_log_activity, _osc_resolve_case_id, _osc_safe_int = _get_osc_helpers()
    row, _ = _osc_exec("SELECT * FROM recurring_expenses WHERE id=%s", (row_id,), fetch="one")
    if not row:
        return jsonify({"ok": False, "error": "not found"}), 404
    payload = request.get_json(silent=True) or {}
    today = date.today()
    start_date = (payload.get("start_date") or f"{today.year}-01-01").strip()
    end_date = (payload.get("end_date") or str(today)).strip()
    try:
        amount = float(payload.get("amount") if "amount" in payload else row.get("amount") or 0)
    except Exception:
        return jsonify({"ok": False, "error": "amount invalid"}), 400
    category = (payload.get("category") or row.get("category") or "").strip()
    sub_type = (payload.get("sub_type") or row.get("sub_type") or "").strip()
    label = (payload.get("description") or row.get("description") or row.get("sub_type") or row.get("category") or "").strip()
    fixed_description = f"[固定] {label}".strip()
    if not category or not fixed_description:
        return jsonify({"ok": False, "error": "recurring expense is incomplete"}), 400
    result, _ = _osc_exec(
        """
        UPDATE case_transactions
           SET amount=%s, category=%s, sub_type=%s
         WHERE type='支出'
           AND date >= %s
           AND date <= %s
           AND COALESCE(category,'')=%s
           AND COALESCE(sub_type,'')=%s
           AND COALESCE(description,'')=%s
        """,
        (amount, category, sub_type or None, start_date, end_date, row.get("category") or "", row.get("sub_type") or "", fixed_description),
        fetch="none",
    )
    updated, _ = _osc_exec(
        """
        SELECT id, case_id, date, type, sub_type, category, description, amount
          FROM case_transactions
         WHERE type='支出'
           AND date >= %s
           AND date <= %s
           AND COALESCE(category,'')=%s
           AND COALESCE(sub_type,'')=%s
           AND COALESCE(description,'')=%s
         ORDER BY date DESC, id DESC
         LIMIT 80
        """,
        (start_date, end_date, category, sub_type, fixed_description),
        fetch="all",
    )
    return jsonify(
        {
            "ok": True,
            "row_id": row_id,
            "start_date": start_date,
            "end_date": end_date,
            "updated_count": (result or {}).get("rowcount", 0),
            "items": updated or [],
        }
    )
