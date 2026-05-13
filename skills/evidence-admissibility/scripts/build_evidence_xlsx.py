#!/usr/bin/env python3
"""
卷證索引證據能力分類腳本（模板）

使用方式：
1. 根據本案資料，修改下方 ========== 可客製化區域 ========== 中的變數
2. 執行腳本即可產出 Excel 檔案

需要客製化的項目：
- INPUT_FILE: 卷證索引文字檔路徑
- OUTPUT_FILE: 輸出 Excel 路徑
- DEFENDANT_NAME: 被告姓名
- section_rules: 卷證索引段落辨識規則
- PROSECUTOR_ADOPTED_PERSONS: 檢察官補充理由書援引的人物
- PERSONS_WITH_TRIAL_TRANSCRIPT: 已於審判中作證的證人
- SECRET_WITNESS_IDS: 秘密證人代號
"""

import re
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================================
# 可客製化區域 — 每案必須根據實際資料修改
# ==========================================================

# 輸入/輸出檔案路徑
INPUT_FILE = '/path/to/卷證索引.txt'
OUTPUT_FILE = '/path/to/output/卷證索引_DEFENDANT_證據能力.xlsx'

# 被告姓名（影響表格標題及被告本人判斷）
DEFENDANT_NAME = '○○○'

# 段落辨識規則
# 格式: (搜尋字串, 段落名稱, 段落類型, 人物ID)
# 段落類型: witness / witness_secret / codefendant / defendant_self / documentary
# 請依卷證索引中出現的順序填寫
section_rules = [
    # 證人
    # ('對於證人○○○', '證人○○○', 'witness', '○○○'),

    # 秘密證人
    # ('對於證人即前信徒A1', '秘密證人A1（前信徒）', 'witness_secret', 'A1'),

    # 同案被告（以證人身份作證的部分）
    # ('同案被告㈠○○○', '同案被告㈠○○○', 'codefendant', '○○○'),

    # 被告本人
    # ('同案被告㈣○○○', '被告○○○（本人）', 'defendant_self', '○○○'),

    # 同案被告的供述（非證人身份）
    # ('被告㈠○○○在警詢、偵查中及本院庭期中之供述', '同案被告㈠○○○（供述）', 'codefendant', '○○○'),

    # 書證/物證（通常在最後）
    ('對於下列證據，有何意見', '書證/物證', 'documentary', 'doc'),
]

# 檢察官補充理由書援引的人物名單
# 從補充理由書的證據清單「壹、供述證據」部分提取
PROSECUTOR_ADOPTED_PERSONS = {
    # '○○○', '△△△',
}

# 已於審判中作證的證人（有審判筆錄 = 已合法調查）
PERSONS_WITH_TRIAL_TRANSCRIPT = {
    # '○○○', '△△△',
}

# 秘密證人代號（其陳述非親自見聞起訴書所載犯罪事實）
SECRET_WITNESS_IDS = {
    # 'A1', 'A3', 'A6',
}

# ==========================================================
# 以下為解析邏輯，通常不需要修改
# ==========================================================

# 讀取卷證索引
with open(INPUT_FILE, 'r', encoding='utf-8') as f:
    text = f.read()

# STEP 1: Split into sections by "審判長問"
parts = re.split(r'(?=審判長問)', text)

# Chinese number items pattern (支援到五十八)
cn_nums = '|'.join([
    '一', '二', '三', '四', '五', '六', '七', '八', '九', '十',
    '十一', '十二', '十三', '十四', '十五', '十六', '十七', '十八', '十九',
    '二十', '二一', '二二', '二三', '二四', '二五', '二六', '二七', '二八', '二九',
    '三十', '三一', '三二', '三三', '三四', '三五', '三六', '三七', '三八', '三九',
    '四十', '四一', '四二', '四三', '四四', '四五', '四六', '四七', '四八', '四九',
    '五十', '五一', '五二', '五三', '五四', '五五', '五六', '五七', '五八',
])

all_items = []
doc_section_counter = 0

for part in parts:
    part_stripped = part.strip()
    if not part_stripped or len(part_stripped) < 20:
        continue

    matched = False
    for search_str, sec_name, sec_type, sec_id in section_rules:
        if search_str in part_stripped:
            if sec_type == 'documentary':
                doc_section_counter += 1
                sec_name = '書證/物證'

            # Extract items by Chinese number
            item_pattern = rf'[　\s]*((?:{cn_nums})、)'
            splits = re.split(item_pattern, part_stripped)

            i = 1
            while i < len(splits):
                if i + 1 < len(splits):
                    item_num = splits[i].strip()
                    item_body = splits[i + 1].strip()

                    # === 文字清理 ===
                    # 消除斷行
                    item_body = re.sub(r'\s*\n\s*', '', item_body)
                    # 消除全形空白和 tab
                    item_body = re.sub(r'[　\t]+', '', item_body)
                    # 合併多餘空白
                    item_body = re.sub(r'  +', ' ', item_body)
                    # 移除程序性文字
                    item_body = re.sub(r'對於下列證據，有何意見？?', '', item_body)
                    item_body = re.sub(r'（提示並告以要旨）', '', item_body)
                    item_body = re.sub(r'提示並告以要旨', '', item_body)
                    item_body = re.sub(r'是否實在？是否出於自由意志？有無意見？?', '', item_body)
                    # 移除供述程序性文字（可根據案件調整）
                    item_body = re.sub(r'在警詢、偵查、本院庭期中、[^之]+之供述', '', item_body)
                    item_body = re.sub(r'在警詢、偵查中及本院庭期中之供述', '', item_body)
                    item_body = re.sub(r'在偵查及本院庭期中之供述', '', item_body)
                    item_body = re.sub(r'在偵查中及本院庭期中之供述', '', item_body)
                    item_body = item_body.strip()

                    all_items.append({
                        'section_name': sec_name,
                        'section_type': sec_type,
                        'section_id': sec_id,
                        'item_text': item_body,
                    })
                    i += 2
                else:
                    i += 1

            matched = True
            break

    if not matched and '對檢察官所提出' in part_stripped:
        continue
    elif not matched and '被告答' in part_stripped:
        continue

print(f'Total parsed items: {len(all_items)}')
type_counts = Counter(s['section_type'] for s in all_items)
for k, v in type_counts.most_common():
    print(f'  {k}: {v}')


# ==========================================================
# STEP 2: 分類函數
# ==========================================================

def classify_record_type(item_text):
    """辨識證據類型（筆錄種類）"""
    t = item_text
    if '警詢筆錄' in t:
        return '警詢筆錄'
    elif '偵訊筆錄' in t or '偵詢筆錄' in t:
        return '偵訊筆錄'
    elif '審判筆錄' in t:
        return '審判筆錄'
    elif '準備程序筆錄' in t:
        return '準備程序筆錄'
    elif '協商程序筆錄' in t:
        return '協商程序筆錄'
    elif '訊問筆錄' in t:
        return '訊問筆錄'
    elif '言詞辯論筆錄' in t:
        return '言詞辯論筆錄'
    elif '勘驗筆錄' in t or '勘驗' in t:
        return '勘驗筆錄'
    elif '搜索票' in t or '搜索扣押' in t:
        return '搜索扣押相關'
    elif '數位採證' in t:
        return '數位採證報告'
    else:
        return '書證'


def get_relationship(sec_type, sec_id):
    """判斷與被告之關係"""
    if sec_type == 'defendant_self':
        return '被告本人'
    elif sec_type == 'codefendant':
        return f'被告以外之人—同案被告{sec_id}'
    elif sec_type in ('witness', 'witness_secret'):
        return f'被告以外之人—證人{sec_id}'
    elif sec_type == 'documentary':
        return '—'
    else:
        return '—'


def is_investigation_stage(item_text):
    """
    判斷筆錄是否屬於偵查階段。
    偵查階段卷宗字號：偵字、相字、他字
    法院階段卷宗字號：國審強處、訴字
    """
    invest_keywords = ['相字', '偵字', '他字']
    return any(kw in item_text for kw in invest_keywords)


def get_admissibility(sec_type, sec_id, record_type, item_text):
    """
    依刑事訴訟法傳聞法則，判斷證據能力意見。
    規則優先序如 SKILL.md 所載。
    """
    # Rule 1: 被告本人 → 不爭執
    if sec_type == 'defendant_self':
        return '不爭執證據能力'

    # Rule 2: 審判筆錄 → 不爭執
    if record_type in ('審判筆錄', '言詞辯論筆錄'):
        return '不爭執證據能力'

    # Rule 3: 法院程序筆錄 → 不爭執
    if record_type in ('準備程序筆錄', '協商程序筆錄'):
        return '不爭執證據能力'

    # Rule 4: 訊問筆錄區分法院階段 vs 偵查階段
    if record_type == '訊問筆錄':
        if is_investigation_stage(item_text):
            pass  # 偵查階段，繼續適用傳聞法則
        else:
            return '不爭執證據能力'  # 法院階段

    # Rule 5: 書證/物證 → 不爭執
    if sec_type == 'documentary':
        return '不爭執證據能力'

    # Rule 6: 被告以外之人的警詢筆錄 → 傳聞
    if record_type == '警詢筆錄' and sec_type != 'defendant_self':
        return '被告以外之人於審判外之陳述，屬傳聞證據，無證據能力（刑事訴訟法第159條第1項）'

    # 偵訊筆錄或偵查階段訊問筆錄
    is_prosecution_record = (record_type == '偵訊筆錄') or \
        (record_type == '訊問筆錄' and is_investigation_stage(item_text))

    # Rule 7: 秘密證人 → 非親自見聞
    if sec_type == 'witness_secret' and sec_id in SECRET_WITNESS_IDS and is_prosecution_record:
        return '所述非親自見聞起訴書所載犯罪事實，屬傳聞證據，無證據能力（刑事訴訟法第159條之1第2項）'

    # Rule 8: 被告以外之人的偵訊筆錄 → 三層判斷
    if is_prosecution_record and sec_type in ('codefendant', 'witness', 'witness_secret'):
        if sec_id in PERSONS_WITH_TRIAL_TRANSCRIPT:
            # 8a: 已合法調查（有審判筆錄）
            return '被告以外之人於審判外向檢察官所為之陳述，屬傳聞證據，不同意作為證據（刑事訴訟法第159條之1第2項）'
        elif sec_id in PROSECUTOR_ADOPTED_PERSONS:
            # 8b: 檢察官援引但未合法調查
            return '經檢察官援引為證據方法，惟未經被告詰問之合法調查，不得作為判斷之依據（刑事訴訟法第155條第2項）'
        else:
            # 8c: 檢察官未援引
            return '未經檢察官於補充理由書援引為證據方法，屬被告以外之人於審判外之陳述，為傳聞證據，不同意作為證據，無證據能力（刑事訴訟法第159條第1項、第159條之1第2項）'

    # Default
    return '不爭執證據能力'


def extract_source(item_text):
    """提取卷頁出處（括號內容），過濾程序性文字"""
    matches = re.findall(r'（([^）]+)）', item_text)
    if matches:
        filtered = [m for m in matches if '提示並告以要旨' not in m]
        cleaned = []
        for m in filtered:
            m = re.sub(r'[；;]?\s*提示並告以要旨\s*', '', m).strip()
            if m:
                cleaned.append(m)
        return '；'.join(cleaned)
    return ''


# ==========================================================
# STEP 3: 產出 Excel
# ==========================================================

wb = openpyxl.Workbook()
ws = wb.active
ws.title = f'卷證索引—{DEFENDANT_NAME}證據能力'

# Styles
header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
cell_font = Font(name='Arial', size=10)
cell_alignment = Alignment(wrap_text=True, vertical='top')
header_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
light_fill = PatternFill('solid', fgColor='F2F7FB')
white_fill = PatternFill('solid', fgColor='FFFFFF')
hearsay_fill = PatternFill('solid', fgColor='FFF2CC')   # 淺黃：無證據能力
secret_fill = PatternFill('solid', fgColor='FCE4EC')    # 淺粉：秘密證人

headers = [
    '編號',
    '證據出處（人/組別）',
    '證據名稱',
    '證據類型',
    '卷頁出處',
    f'與被告{DEFENDANT_NAME}之關係',
    '證據能力意見',
]
col_widths = [6, 28, 55, 14, 45, 28, 55]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

# Write data rows
for row_idx, item in enumerate(all_items, 2):
    record_type = classify_record_type(item['item_text'])
    relationship = get_relationship(item['section_type'], item['section_id'])
    admissibility = get_admissibility(
        item['section_type'], item['section_id'], record_type, item['item_text']
    )
    source = extract_source(item['item_text'])

    row_data = [
        row_idx - 1,
        item['section_name'],
        item['item_text'],
        record_type,
        source,
        relationship,
        admissibility,
    ]

    # Color coding
    if '無證據能力' in admissibility:
        fill = hearsay_fill
    elif row_idx % 2 == 0:
        fill = light_fill
    else:
        fill = white_fill

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border
        cell.fill = fill

# Freeze header row & auto-filter
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:G{len(all_items) + 1}'

wb.save(OUTPUT_FILE)

# Summary
print(f'\nSaved to {OUTPUT_FILE}')
print(f'Total rows: {len(all_items)}')

admissibility_counts = Counter()
for item in all_items:
    rt = classify_record_type(item['item_text'])
    adm = get_admissibility(item['section_type'], item['section_id'], rt, item['item_text'])
    admissibility_counts[adm] += 1

print('\n=== 證據能力意見分佈 ===')
for k, v in admissibility_counts.most_common():
    print(f'  [{v}] {k}')
