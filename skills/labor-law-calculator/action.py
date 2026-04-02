#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
labor-law-calculator/action.py

台灣勞動基準法計算技能
- 加班費（平日/休息日/例假日/國定假日，含各修法版本）
- 特休假天數
- 資遣費（舊制/新制/混合制）
- 試算自我驗證
- 支援 Excel/Google Sheets URL 自動代算
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib import request, parse

# ─── Constants ───────────────────────────────────────────────────────────────

# 法定基本時薪（歷年調整）
_MIN_HOURLY_WAGE_HISTORY = [
    (date(2026, 1, 1), 190),   # 2026 時薪（暫同 2025）
    (date(2025, 1, 1), 190),   # 2025 時薪 190 元
    (date(2024, 1, 1), 183),
    (date(2023, 1, 1), 176),
    (date(2022, 1, 1), 168),
    (date(2021, 1, 1), 160),
    (date(2020, 1, 1), 158),
    (date(2019, 1, 1), 150),
    (date(2018, 1, 1), 140),
    (date(2017, 1, 1), 133),
    (date(2016, 1, 1), 126),
    (date(2015, 7, 1), 120),
]

# 法定基本月薪（歷年調整）
_MIN_MONTHLY_WAGE_HISTORY = [
    (date(2026, 1, 1), 28590),  # 2026 基本工資（暫同 2025）
    (date(2025, 1, 1), 28590),  # 2025 基本工資 28,590 元
    (date(2024, 1, 1), 27470),
    (date(2023, 1, 1), 26400),
    (date(2022, 1, 1), 25250),
    (date(2021, 1, 1), 24000),
    (date(2020, 1, 1), 23800),
    (date(2019, 1, 1), 23100),
    (date(2018, 1, 1), 22000),
    (date(2017, 1, 1), 21009),
    (date(2016, 1, 1), 20008),
    (date(2015, 7, 1), 20008),
]

# 修法版本分界
_AMENDMENT_2016 = date(2016, 1, 1)   # 一例一休前：週 48h
_AMENDMENT_2018 = date(2018, 3, 1)   # 彈性工時修正施行

# 新制資遣費起算日
_NEW_SEVERANCE_START = date(2005, 7, 1)

# ─── 工資組成結構 ─────────────────────────────────────────────────────────────
#
# 勞基法第 2 條第 3 款：工資 = 勞工因工作而獲得之報酬，包括薪金及其他任何名義
# 之「經常性給與」（不問名目）。
# 勞基法施行細則第 10 條：下列各款不視為工資（非經常性給與）：
#   紅利、年終獎金、競賽獎金、三節獎金（端午/中秋/年節）、醫療補助費、
#   教育補助費、差旅費、交際費、作業用品代金等。
#
# 本計算器採用的判斷原則：
#   ✅ 算入（每月固定給付）：本薪、職務加給、技術加給、伙食津貼（每月固定）、
#                          全勤獎金（固定）、交通津貼（固定）、其他固定津貼
#   ❌ 不算入（非經常性）：年終獎金、三節獎金、績效獎金、差旅費報銷、健保勞保費

_REGULAR_WAGE_MEAL_EXEMPT = 3000  # 2024起每月免稅額（僅供說明，不影響工資認定）

@dataclass
class WageComponents:
    """月工資組成（用於計算加班費時薪基礎）"""
    base: float             # 本薪/底薪
    meal: float = 0.0       # 伙食津貼/午餐加給（每月固定給付者）
    attendance_bonus: float = 0.0  # 全勤獎金（每月固定）
    transport: float = 0.0  # 交通津貼（每月固定）
    duty_allowance: float = 0.0    # 職務加給/技術加給
    other_fixed: float = 0.0       # 其他每月固定津貼

    @property
    def total(self) -> float:
        return (self.base + self.meal + self.attendance_bonus
                + self.transport + self.duty_allowance + self.other_fixed)

    def breakdown_str(self) -> str:
        parts = [f"本薪 {self.base:,.0f}"]
        if self.meal:
            parts.append(f"伙食津貼 {self.meal:,.0f}")
        if self.attendance_bonus:
            parts.append(f"全勤獎金 {self.attendance_bonus:,.0f}")
        if self.transport:
            parts.append(f"交通津貼 {self.transport:,.0f}")
        if self.duty_allowance:
            parts.append(f"職務加給 {self.duty_allowance:,.0f}")
        if self.other_fixed:
            parts.append(f"其他固定津貼 {self.other_fixed:,.0f}")
        return " + ".join(parts) + f" = **{self.total:,.0f} 元**（經常性薪資）"


def _components_from_wage(w) -> WageComponents:
    """接受 float（視為純本薪）或 WageComponents 物件。"""
    if isinstance(w, WageComponents):
        return w
    return WageComponents(base=float(w))


# ─── Data Classes ────────────────────────────────────────────────────────────

@dataclass
class OvertimeResult:
    base_hourly: float
    day_type: str          # 平日/休息日/例假日/國定假日
    hours: float
    extra_pay: float       # 加班加給部分
    total_pay: float       # 含正常工資部分
    breakdown: List[str]
    law_basis: str
    validation_ok: bool
    validation_note: str

@dataclass
class AnnualLeaveResult:
    seniority_months: int
    seniority_years: float
    annual_days: int
    breakdown: str
    law_basis: str

@dataclass
class SeveranceResult:
    system: str            # old/new/mixed
    avg_wage: float
    old_years: float
    new_years: float
    old_amount: float
    new_amount: float
    total_amount: float
    breakdown: List[str]
    law_basis: str
    validation_ok: bool

# ─── Helpers ─────────────────────────────────────────────────────────────────

def _round2(v: float) -> float:
    return float(Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

def _round0(v: float) -> int:
    return int(Decimal(str(v)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))

def _min_hourly(ref_date: date) -> int:
    for d, w in _MIN_HOURLY_WAGE_HISTORY:
        if ref_date >= d:
            return w
    return 120

def _min_monthly(ref_date: date) -> int:
    for d, w in _MIN_MONTHLY_WAGE_HISTORY:
        if ref_date >= d:
            return w
    return 20008

def _hourly_from_monthly(monthly) -> float:
    """經常性薪資（月）÷ 30 ÷ 8 = 時薪（勞基法施行細則第 11 條）"""
    comp = _components_from_wage(monthly)
    return comp.total / 30.0 / 8.0

def _parse_date(s: str) -> Optional[date]:
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

# ─── 加班費計算 ───────────────────────────────────────────────────────────────

def calc_overtime(
    monthly_wage,             # float 或 WageComponents
    hours: float,
    day_type: str,            # 平日/休息日/例假日/國定假日
    ref_date: Optional[date] = None,
) -> OvertimeResult:
    """
    計算單日加班費。

    修法版本邏輯：
    - 2018-03-01 後：目前現行制
    - 2016-01-01～2018-02-28：一例一休期間（與現行相同，補充說明）
    - 2016-01-01 前：舊制（週 48h，無休息日概念）

    加班費規則：
    - 平日延長工時：前 2h × 4/3，第 3h 起 × 5/3
    - 休息日（2016～2018 一例一休版）：前 2h 加給 × 4/3，3-8h × 5/3，9-12h × 8/3
      ※ 出勤即以 4h 計，超過 4h 以 8h 計，超過 8h 以 12h 計
    - 休息日（2018/3/1 後現行）：前 2h 加給 × 4/3，第 3h 起 × 5/3
      ※ 按實際出勤時數計算，無最低計費時數
    - 例假日：強制休息，出勤須支付額外 1 日工資（×2 計算）
    - 國定假日：出勤另補 1 日工資（×2 計算）
    """
    if ref_date is None:
        ref_date = date.today()

    comp = _components_from_wage(monthly_wage)
    hourly = comp.total / 30.0 / 8.0
    day_type = day_type.strip()

    breakdown: List[str] = [
        f"經常性薪資：{comp.breakdown_str()}",
        f"時薪 = {comp.total:,.0f} ÷ 30 ÷ 8 = {hourly:.4f} 元",
    ]

    # ── 平日延長工時 ──
    if day_type == "平日":
        if ref_date < _AMENDMENT_2016:
            law = "勞基法第 24 條（2016 年修正前，舊制週 48h）"
            note = "舊制：前 2h ×4/3，第 3h 起 ×5/3（同現行）"
        elif ref_date < _AMENDMENT_2018:
            law = "勞基法第 24 條（2016 一例一休版）"
            note = "一例一休：前 2h ×4/3，第 3h 起 ×5/3"
        else:
            law = "勞基法第 24 條第 1 項（2018 年現行）"
            note = "現行：前 2h ×4/3，第 3h 起 ×5/3"

        h1 = min(hours, 2.0)
        h2 = max(hours - 2.0, 0.0)
        extra = _round2(hourly * h1 * (4/3 - 1) + hourly * h2 * (5/3 - 1))
        total = _round2(hourly * hours + extra)
        breakdown += [
            f"前 {h1:.1f}h × 時薪 × (4/3-1) = {hourly*h1*(4/3-1):.2f}",
            f"後 {h2:.1f}h × 時薪 × (5/3-1) = {hourly*h2*(5/3-1):.2f}",
            f"加班加給合計 = {extra:.2f}，當日應得薪資 = {total:.2f}",
            note,
        ]
        val_ok = abs(extra - _round2(hourly * h1 * 1/3 + hourly * h2 * 2/3)) < 0.02
        val_note = "自我驗證：加給 = h1×1/3 + h2×2/3 ✓" if val_ok else "驗證異常"
        return OvertimeResult(hourly, day_type, hours, extra, total, breakdown, law, val_ok, val_note)

    # ── 休息日 ──
    elif day_type == "休息日":
        if ref_date < _AMENDMENT_2016:
            law = "勞基法第 24 條（2016 年前無休息日規定，依例假日處理）"
            # 2016 前無休息日概念，改以平日加班計算
            return calc_overtime(monthly_wage, hours, "平日", ref_date)

        # 正常日薪（月薪已含，不另計入加班費）
        normal_day_pay = hourly * 8.0

        if ref_date < _AMENDMENT_2018:
            # ── 2016～2018 一例一休版：最低計費時數 + 三級費率 ──
            law = "勞基法第 24 條第 2 項（2016 一例一休版）"
            # 出勤即以 4h 計；超過 4h 以 8h 計；超過 8h 以 12h 計
            billed_hours = 4.0 if hours <= 4.0 else (8.0 if hours <= 8.0 else min(hours, 12.0))
            bh = billed_hours
            h1 = min(bh, 2.0)
            h2 = min(max(bh - 2.0, 0.0), 6.0)   # 3-8h
            h3 = max(bh - 8.0, 0.0)              # 9-12h
            extra = _round2(hourly * h1 * 4/3 + hourly * h2 * 5/3 + hourly * h3 * 8/3)
            total = _round2(normal_day_pay + extra)
            breakdown += [
                f"實出勤 {hours:.1f}h → 計費時數 {billed_hours:.1f}h（出勤≤4h以4h計，≤8h以8h計）",
                f"前 {h1:.1f}h × {hourly:.2f} × 4/3 = {hourly*h1*4/3:.2f}",
                f"3-8h 段 {h2:.1f}h × {hourly:.2f} × 5/3 = {hourly*h2*5/3:.2f}",
                f"9-12h 段 {h3:.1f}h × {hourly:.2f} × 8/3 = {hourly*h3*8/3:.2f}",
                f"加給合計 {extra:.2f}（另加正常日薪 {normal_day_pay:.2f}）",
                f"當日應得薪資 = {total:.2f}",
            ]
            val_extra = _round2(hourly*h1*4/3 + hourly*h2*5/3 + hourly*h3*8/3)
            val_ok = abs(extra - val_extra) < 0.02
            val_note = "自我驗證：逐段累加一致 ✓" if val_ok else "驗證異常"
        else:
            # ── 2018/3/1 後現行版：按實際時數、二級費率 ──
            law = "勞基法第 24 條第 2 項（2018 年現行）"
            # 2018 修法移除最低計費時數，按實際出勤時數計算
            h1 = min(hours, 2.0)
            h2 = max(hours - 2.0, 0.0)
            extra = _round2(hourly * h1 * 4/3 + hourly * h2 * 5/3)
            total = _round2(normal_day_pay + extra)
            breakdown += [
                f"休息日出勤 {hours:.1f}h（2018 修法後按實際時數計算，無最低計費）",
                f"前 {h1:.1f}h × {hourly:.2f} × 4/3 = {hourly*h1*4/3:.2f}",
                f"第 3h 起 {h2:.1f}h × {hourly:.2f} × 5/3 = {hourly*h2*5/3:.2f}",
                f"加給合計 {extra:.2f}（另加正常日薪 {normal_day_pay:.2f}）",
                f"當日應得薪資 = {total:.2f}",
            ]
            val_extra = _round2(hourly*h1*4/3 + hourly*h2*5/3)
            val_ok = abs(extra - val_extra) < 0.02
            val_note = "自我驗證：逐段累加一致 ✓" if val_ok else "驗證異常"

        return OvertimeResult(hourly, day_type, hours, extra, total, breakdown, law, val_ok, val_note)

    # ── 例假日 ──
    elif day_type == "例假日":
        law = "勞基法第 39 條（例假日出勤加倍給付）"
        if hours > 8:
            hours = 8.0
            breakdown.append("例假日最多以 8h 計")
        extra = _round2(hourly * hours)   # 加倍 = 再給 1 日
        total = _round2(hourly * hours * 2)
        breakdown += [
            f"例假日出勤 {hours:.1f}h，除正常薪資外另加 1 倍",
            f"額外加給 = {extra:.2f}，當日薪資合計 = {total:.2f}",
        ]
        val_ok = abs(total - hourly * hours * 2) < 0.02
        return OvertimeResult(hourly, day_type, hours, extra, total, breakdown, law, val_ok, "×2 驗證 ✓" if val_ok else "驗證異常")

    # ── 國定假日 ──
    elif day_type in ("國定假日", "紀念日", "節日"):
        day_type = "國定假日"
        law = "勞基法第 39 條（國定假日出勤加倍給付）"
        if hours > 8:
            hours = 8.0
            breakdown.append("國定假日最多以 8h 計")
        extra = _round2(hourly * hours)
        total = _round2(hourly * hours * 2)
        breakdown += [
            f"國定假日出勤 {hours:.1f}h，除正常薪資外另加 1 倍",
            f"額外加給 = {extra:.2f}，當日薪資合計 = {total:.2f}",
        ]
        val_ok = abs(total - hourly * hours * 2) < 0.02
        return OvertimeResult(hourly, day_type, hours, extra, total, breakdown, law, val_ok, "×2 驗證 ✓" if val_ok else "驗證異常")

    else:
        raise ValueError(f"未知假別：{day_type}，請輸入 平日/休息日/例假日/國定假日")


# ─── 特休假計算 ───────────────────────────────────────────────────────────────

def calc_annual_leave(
    start_date: date,
    ref_date: Optional[date] = None,
) -> AnnualLeaveResult:
    """
    計算特休假天數。

    勞基法第 38 條（2017 年修正後現行）：
    6個月以上未滿1年 → 3天
    1年以上未滿2年  → 7天
    2年以上未滿3年  → 10天
    3年以上未滿5年  → 14天
    5年以上未滿10年 → 15天
    10年以上        → 每滿一年加 1 天，最多 30 天
    """
    if ref_date is None:
        ref_date = date.today()

    # 計算服務月數（精確到天）
    delta_days = (ref_date - start_date).days
    months = delta_days / 30.4375
    years = delta_days / 365.25

    if months < 6:
        days = 0
        desc = f"服務未滿 6 個月（{months:.1f} 個月），尚無特休假"
    elif months < 12:
        days = 3
        desc = f"服務 6 個月以上未滿 1 年（{months:.1f} 個月）→ 3 天"
    elif years < 2:
        days = 7
        desc = f"服務 1 年以上未滿 2 年（{years:.2f} 年）→ 7 天"
    elif years < 3:
        days = 10
        desc = f"服務 2 年以上未滿 3 年（{years:.2f} 年）→ 10 天"
    elif years < 5:
        days = 14
        desc = f"服務 3 年以上未滿 5 年（{years:.2f} 年）→ 14 天"
    elif years < 10:
        days = 15
        desc = f"服務 5 年以上未滿 10 年（{years:.2f} 年）→ 15 天"
    else:
        extra = min(int(years) - 10, 15)
        days = min(15 + extra, 30)
        desc = f"服務 {years:.2f} 年（10 年以上）→ 15+{extra}={days} 天（上限 30 天）"

    return AnnualLeaveResult(
        seniority_months=int(months),
        seniority_years=round(years, 2),
        annual_days=days,
        breakdown=desc,
        law_basis="勞基法第 38 條（2017 年 3 月 1 日修正施行）",
    )


# ─── 資遣費計算 ───────────────────────────────────────────────────────────────

def calc_severance(
    avg_wage: float,
    hire_date: date,
    terminate_date: Optional[date] = None,
    labor_pension_choice: str = "auto",   # old/new/auto
) -> SeveranceResult:
    """
    計算資遣費。

    舊制（勞基法第 17 條）：每滿 1 年給 1 個月平均工資，最高 6 個月。
    新制（勞工退休金條例第 12 條，2005/7/1 起）：每滿 1 年給 0.5 個月，最高 6 個月。
    混合制：2005/7/1 前年資用舊制，之後用新制（各自上限 6 個月）。

    labor_pension_choice:
      'old'  → 全部舊制（留在舊制者）
      'new'  → 全部新制（2005/7/1 後到職者）
      'auto' → 2005/7/1 前到職自動混合計算
    """
    if terminate_date is None:
        terminate_date = date.today()

    breakdown: List[str] = []
    old_years = new_years = 0.0
    old_amount = new_amount = 0.0

    total_days = (terminate_date - hire_date).days
    total_years = total_days / 365.25

    if labor_pension_choice == "old":
        # 全部視同舊制（明確選擇適用勞退舊制）
        old_years = total_years
        new_years = 0.0
        # 舊制：每滿 1 年 1 個月，最高 6 個月；未滿 1 年依比例
        capped_years = min(old_years, 6.0)
        old_amount = _round0(avg_wage * capped_years)
        breakdown.append(f"舊制年資 {old_years:.3f} 年 → min({old_years:.3f},6) × {avg_wage:.0f} = {old_amount:.0f} 元")
        total = old_amount
        system = "old"

    elif labor_pension_choice == "new" or (labor_pension_choice == "auto" and hire_date >= _NEW_SEVERANCE_START):
        # 全部新制
        new_years = total_years
        old_years = 0.0
        capped_years = min(new_years, 12.0)  # 0.5 × 12 = 6 個月上限
        new_amount = _round0(avg_wage * 0.5 * capped_years)
        breakdown.append(f"新制年資 {new_years:.3f} 年 → min({new_years:.3f},12) × 0.5 × {avg_wage:.0f} = {new_amount:.0f} 元")
        total = new_amount
        system = "new"

    else:
        # auto = 混合制（2005/7/1 前到職）
        system = "mixed"
        pivot = _NEW_SEVERANCE_START

        if hire_date < pivot:
            old_days = (pivot - hire_date).days
            new_days = max((terminate_date - pivot).days, 0)
        else:
            old_days = 0
            new_days = total_days

        old_years = old_days / 365.25
        new_years = new_days / 365.25

        # 舊制部分：每滿 1 年 1 月，上限 6 個月
        old_capped = min(old_years, 6.0)
        old_amount = _round0(avg_wage * old_capped)
        breakdown.append(f"舊制（迄 2005/7/1）{old_years:.3f} 年 → {old_capped:.3f} × {avg_wage:.0f} = {old_amount:.0f} 元")

        # 新制部分：每滿 1 年 0.5 月，上限 6 個月
        new_capped = min(new_years, 12.0)
        new_amount = _round0(avg_wage * 0.5 * new_capped)
        breakdown.append(f"新制（2005/7/1 起）{new_years:.3f} 年 → {new_capped:.3f} × 0.5 × {avg_wage:.0f} = {new_amount:.0f} 元")

        total = old_amount + new_amount

    law = "勞基法第 17 條（舊制）/ 勞工退休金條例第 12 條（新制）"

    # 自我驗證
    val_check = 0.0
    if system == "old":
        val_check = avg_wage * min(old_years, 6.0)
    elif system == "new":
        val_check = avg_wage * 0.5 * min(new_years, 12.0)
    else:
        val_check = avg_wage * min(old_years, 6.0) + avg_wage * 0.5 * min(new_years, 12.0)
    val_ok = abs(total - _round0(val_check)) <= 1

    return SeveranceResult(
        system=system,
        avg_wage=avg_wage,
        old_years=round(old_years, 3),
        new_years=round(new_years, 3),
        old_amount=old_amount,
        new_amount=new_amount,
        total_amount=total,
        breakdown=breakdown,
        law_basis=law,
        validation_ok=val_ok,
    )


# ─── Excel / Google Sheets 自動代算 ──────────────────────────────────────────

def _fetch_sheet_data(url: str) -> Optional[Dict]:
    """
    嘗試從 Google Sheets 匯出 CSV，解析欄位。
    支援公開分享的試算表，格式：
    A欄=月薪, B欄=日期類型, C欄=加班時數, D欄=到職日
    """
    # 轉換成 CSV 匯出 URL
    sheet_id = None
    if "docs.google.com/spreadsheets" in url:
        m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
        if m:
            sheet_id = m.group(1)
            gid_m = re.search(r"gid=(\d+)", url)
            gid = gid_m.group(1) if gid_m else "0"
            csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    else:
        return None

    try:
        ctx = __import__("ssl").create_default_context()
        with request.urlopen(csv_url, timeout=10) as resp:
            raw = resp.read().decode("utf-8-sig")
    except Exception as e:
        return {"error": str(e)}

    import csv
    import io
    rows = list(csv.reader(io.StringIO(raw)))
    if not rows:
        return {"error": "空試算表"}
    return {"rows": rows, "sheet_id": sheet_id}


def calc_from_sheet(url: str, ref_date: Optional[date] = None) -> str:
    """
    從 Google Sheets URL 自動代算加班費。
    試算表格式（第一列為標題，第二列起為資料）：
    月薪 | 假別（平日/休息日/例假日/國定假日）| 加班時數 | 到職日（YYYY-MM-DD，選填）
    """
    result = _fetch_sheet_data(url)
    if result is None:
        return "不支援此試算表格式，目前支援 Google Sheets 公開試算表。"
    if "error" in result:
        return f"讀取試算表失敗：{result['error']}"

    rows = result.get("rows", [])
    if len(rows) < 2:
        return "試算表沒有資料列（第 2 列起為資料）。"

    lines = ["📊 **加班費自動試算結果**\n"]
    for i, row in enumerate(rows[1:], start=2):
        if len(row) < 3:
            continue
        try:
            monthly = float(str(row[0]).replace(",", "").replace("$", ""))
            day_type = row[1].strip()
            hours = float(row[2])
        except (ValueError, IndexError):
            lines.append(f"第 {i} 列資料格式錯誤，跳過。")
            continue

        hire_date = None
        if len(row) >= 4 and row[3].strip():
            hire_date = _parse_date(row[3].strip())

        ot = calc_overtime(monthly, hours, day_type, ref_date)
        lines.append(f"**第 {i} 列**：月薪 {monthly:.0f}，{day_type} {hours}h")
        lines.append(f"  加班加給：{ot.extra_pay:.0f} 元，當日合計：{ot.total_pay:.0f} 元")
        lines.append(f"  法條：{ot.law_basis}")

        if hire_date:
            al = calc_annual_leave(hire_date, ref_date)
            lines.append(f"  特休假：{al.annual_days} 天（{al.breakdown}）")

    return "\n".join(lines)


# ─── 格式化輸出 ───────────────────────────────────────────────────────────────

def _fmt_overtime(r: OvertimeResult) -> str:
    lines = [
        f"⏱ **加班費試算（{r.day_type}）**",
        f"時薪：{r.base_hourly:.4f} 元",
        f"加班時數：{r.hours:.1f}h",
        f"加班加給：**{r.extra_pay:.0f} 元**",
        f"當日薪資合計：**{r.total_pay:.0f} 元**",
        "",
        "計算明細：",
    ]
    lines += [f"  {b}" for b in r.breakdown]
    lines += [
        f"\n法條依據：{r.law_basis}",
        f"自我驗證：{'✅ ' + r.validation_note if r.validation_ok else '⚠️ ' + r.validation_note}",
    ]
    return "\n".join(lines)


def _fmt_annual_leave(r: AnnualLeaveResult) -> str:
    return (
        f"🏖 **特休假試算**\n"
        f"服務年資：{r.seniority_years:.2f} 年（{r.seniority_months} 個月）\n"
        f"特休天數：**{r.annual_days} 天**\n"
        f"說明：{r.breakdown}\n"
        f"法條依據：{r.law_basis}"
    )


def _fmt_severance(r: SeveranceResult) -> str:
    sys_label = {"old": "舊制", "new": "新制", "mixed": "混合制"}.get(r.system, r.system)
    lines = [
        f"💰 **資遣費試算（{sys_label}）**",
        f"平均工資：{r.avg_wage:.0f} 元",
    ]
    if r.old_years > 0:
        lines.append(f"舊制年資：{r.old_years:.3f} 年 → {r.old_amount:.0f} 元")
    if r.new_years > 0:
        lines.append(f"新制年資：{r.new_years:.3f} 年 → {r.new_amount:.0f} 元")
    lines += [
        f"資遣費合計：**{r.total_amount:.0f} 元**",
        "",
        "計算明細：",
    ]
    lines += [f"  {b}" for b in r.breakdown]
    lines += [
        f"\n法條依據：{r.law_basis}",
        f"自我驗證：{'✅ 驗算一致' if r.validation_ok else '⚠️ 驗算異常，請人工複核'}",
    ]
    return "\n".join(lines)


# ─── 本地 Excel 出席明細紀錄表解析 ──────────────────────────────────────────────

# 台灣民國年 → 西元年
def _roc_to_ad(roc_year: int) -> int:
    return roc_year + 1911

# 打卡系統 HHMM 格式（0123 → 83 分鐘）
def _hhmm_to_minutes(s: str) -> int:
    s = str(s).strip().zfill(4)
    try:
        return int(s[:2]) * 60 + int(s[2:])
    except Exception:
        return 0

# 台灣國定假日（西元）
_TW_HOLIDAYS: Dict[Tuple[int, int, int], str] = {}

def _build_tw_holidays(year: int) -> Dict[Tuple[int, int, int], str]:
    """產生指定年份已知固定國定假日（不含農曆浮動節日）。"""
    h: Dict[Tuple[int, int, int], str] = {
        (year, 1,  1): "元旦",
        (year, 2, 28): "和平紀念日",
        (year, 4,  4): "兒童節/婦女節",
        (year, 4,  5): "清明節(4/5)",
        (year, 5,  1): "勞動節",
        (year, 10, 10): "國慶日",
        (year, 12, 25): "行憲紀念日",
    }
    return h

def _get_day_type_from_weekday(weekday_char: str) -> str:
    """從民國日期行星期字符判斷假別。日=例假日 六=休息日 一~五=平日"""
    if weekday_char == "日":
        return "例假日"
    elif weekday_char == "六":
        return "休息日"
    else:
        return "平日"

@dataclass
class DailyOTRecord:
    date_str: str          # e.g. "110/12/01"
    ad_date: date
    weekday: str           # 一二三四五六日
    day_type: str          # 平日/休息日/例假日/國定假日/停班停課
    pre_ot_min: int        # 上班前加班（分鐘）
    post_ot_min: int       # 申報加班（分鐘）
    total_ot_min: int      # pre+post
    source: str            # excel/pdf
    note: str = ""

def _parse_attendance_excel(path: str) -> List[DailyOTRecord]:
    """
    解析出席明細紀錄表 xlsx（打卡系統匯出格式）。
    每行格式：
      YYYY/MM/DD(星期) 應上班 應下班 實上班 實下班 ... 上班前加班 申報加班 0000
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("需要 openpyxl：pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    records: List[DailyOTRecord] = []

    # Regex: 日期行
    # e.g. "       110/12/01(三)    08:30  17:30  08:01  20:44  ...  0029  0314  0000"
    ROW_RE = re.compile(
        r"(\d{3}/\d{2}/\d{2})\(([一二三四五六日])\)"   # date + weekday
        r"(?:.*?(\d{4})\s+(\d{4})\s+(\d{4})\s*$)?"     # pre_ot, post_ot, zero  (trailing 3 groups optional)
    )
    OT_RE = re.compile(r"(\d{4})\s+(\d{4})\s+\d{4}\s*$")  # last 3 HHMM groups at EOL

    for shname in wb.sheetnames:
        ws = wb[shname]
        for row in ws.iter_rows(values_only=True):
            raw = str(row[0]) if row[0] is not None else ""
            raw = raw.strip()
            if not raw:
                continue

            m = ROW_RE.search(raw)
            if not m:
                continue

            date_str = m.group(1)   # "110/12/01"
            wday = m.group(2)       # "三"

            # Parse AD date
            parts = date_str.split("/")
            try:
                ad_year = _roc_to_ad(int(parts[0]))
                ad_date = date(ad_year, int(parts[1]), int(parts[2]))
            except Exception:
                continue

            # Extract OT columns (last 3 HHMM groups)
            ot_m = OT_RE.search(raw)
            if not ot_m:
                pre_min = post_min = 0
            else:
                pre_min = _hhmm_to_minutes(ot_m.group(1))
                post_min = _hhmm_to_minutes(ot_m.group(2))

            if pre_min == 0 and post_min == 0:
                continue  # 無加班紀錄，跳過

            # Determine day type
            day_type = _get_day_type_from_weekday(wday)

            # Override for known national holidays
            hols = _build_tw_holidays(ad_date.year)
            hol_key = (ad_date.year, ad_date.month, ad_date.day)
            if hol_key in hols:
                day_type = "國定假日"

            records.append(DailyOTRecord(
                date_str=date_str,
                ad_date=ad_date,
                weekday=wday,
                day_type=day_type,
                pre_ot_min=pre_min,
                post_ot_min=post_min,
                total_ot_min=pre_min + post_min,
                source="excel",
                note="",
            ))

    return sorted(records, key=lambda r: r.ad_date)


def _parse_holiday_pdf(path: str) -> List[DailyOTRecord]:
    """
    解析假日加班紀錄 PDF（手寫格式）。
    每頁格式：
      日期 時間 備註
      YYYY/MM/DD(星期) HH:MM-HH:MM  ...備註...
      或
      YYYY.MM.DD(星期) HH:MM  ...備註...
    """
    try:
        import fitz
    except ImportError:
        raise RuntimeError("需要 PyMuPDF：pip install pymupdf")

    doc = fitz.open(path)
    try:
        return _parse_ot_pdf_pages(doc)
    finally:
        doc.close()


def _parse_ot_pdf_pages(doc) -> "List[DailyOTRecord]":
    """Internal: parse opened fitz.Document pages into OT records."""
    records: List[DailyOTRecord] = []

    DATE_RE = re.compile(
        r"(\d{3,4})[./](\d{1,2})[./](\d{1,2})\(([一二三四五六日])\)"
    )
    TIME_RANGE_RE = re.compile(r"(\d{1,2}):(\d{2})\s*[-~～]\s*(\d{1,2}):(\d{2})")
    TIME_SINGLE_RE = re.compile(r"(\d{1,2}):(\d{2})")

    for pg in doc:
        text = pg.get_text()
        dm = DATE_RE.search(text)
        if not dm:
            continue

        roc_or_ad_year = int(dm.group(1))
        month = int(dm.group(2))
        day = int(dm.group(3))
        wday = dm.group(4)
        # Determine if ROC or AD year
        ad_year = roc_or_ad_year if roc_or_ad_year > 1911 else _roc_to_ad(roc_or_ad_year)
        try:
            ad_date = date(ad_year, month, day)
        except Exception:
            continue

        date_str = f"{roc_or_ad_year}/{month:02d}/{day:02d}"

        # Parse duration
        tr_m = TIME_RANGE_RE.search(text)
        if tr_m:
            h1, m1, h2, m2 = int(tr_m.group(1)), int(tr_m.group(2)), int(tr_m.group(3)), int(tr_m.group(4))
            total_min = max((h2 * 60 + m2) - (h1 * 60 + m1), 0)
        else:
            ts_m = TIME_SINGLE_RE.search(text)
            total_min = 60 if ts_m else 0   # 單一時間點視為1小時

        if total_min <= 0:
            continue

        day_type = _get_day_type_from_weekday(wday)
        # 颱風停班停課 → 停班停課
        note = text.replace("\n", " ").strip()
        if any(k in note for k in ["颱風", "停班停課", "天災"]):
            day_type = "停班停課"

        # 國定假日
        hols = _build_tw_holidays(ad_date.year)
        if (ad_date.year, month, day) in hols:
            day_type = "國定假日"

        records.append(DailyOTRecord(
            date_str=date_str,
            ad_date=ad_date,
            weekday=wday,
            day_type=day_type,
            pre_ot_min=0,
            post_ot_min=total_min,
            total_ot_min=total_min,
            source="pdf",
            note=note[:80],
        ))

    return sorted(records, key=lambda r: r.ad_date)


def _calc_ot_pay_for_record(rec: DailyOTRecord, monthly_wage) -> float:
    """
    依假別計算單日加班費（加給部分，不含正常工資）。
    停班停課 → 出勤加倍（同例假日，雙倍計算）。
    monthly_wage: float 或 WageComponents（以經常性薪資計算時薪）
    """
    hourly = _hourly_from_monthly(monthly_wage)
    ot_hours = rec.total_ot_min / 60.0

    if rec.day_type in ("例假日", "國定假日", "停班停課"):
        # 加倍：整日（出勤時數）× 時薪
        return _round2(hourly * ot_hours)  # 額外加給部分（original already paid）

    elif rec.day_type == "休息日":
        # 2018/3/1 後現行：按實際時數、二級費率（無最低計費）
        if rec.ad_date >= _AMENDMENT_2018:
            h1 = min(ot_hours, 2.0)
            h2 = max(ot_hours - 2.0, 0.0)
            return _round2(hourly * h1 * 4/3 + hourly * h2 * 5/3)
        else:
            # 2016～2018 一例一休版：最低計費時數 + 三級費率
            billed_h = 4.0 if ot_hours <= 4.0 else (8.0 if ot_hours <= 8.0 else min(ot_hours, 12.0))
            h1 = min(billed_h, 2.0)
            h2 = min(max(billed_h - 2.0, 0.0), 6.0)
            h3 = max(billed_h - 8.0, 0.0)
            return _round2(hourly * h1 * 4/3 + hourly * h2 * 5/3 + hourly * h3 * 8/3)

    else:  # 平日
        h1 = min(ot_hours, 2.0)
        h2 = max(ot_hours - 2.0, 0.0)
        return _round2(hourly * h1 * 1/3 + hourly * h2 * 2/3)


def calc_case_overtime(
    files: List[str],
    monthly_wage,                              # float 或 WageComponents
    monthly_wage_by_year: Optional[Dict[int, Any]] = None,  # {year: float 或 WageComponents}
) -> str:
    """
    計算多個出席 Excel + 假日 PDF 的總加班費。

    Args:
        files: 檔案路徑清單（xlsx 或 pdf）
        monthly_wage: 預設月薪（float）或 WageComponents（含各津貼）
        monthly_wage_by_year: {西元年: float 或 WageComponents} 覆蓋特定年份
    """
    all_records: List[DailyOTRecord] = []
    errors: List[str] = []

    for f in files:
        ext = Path(f).suffix.lower()
        try:
            if ext in (".xlsx", ".xls"):
                recs = _parse_attendance_excel(f)
            elif ext == ".pdf":
                recs = _parse_holiday_pdf(f)
            else:
                errors.append(f"⚠️ 不支援格式：{Path(f).name}")
                continue
            all_records.extend(recs)
        except Exception as e:
            errors.append(f"⚠️ 讀取失敗 {Path(f).name}：{e}")

    if not all_records:
        err_msg = "\n".join(errors) if errors else "無可解析的加班紀錄。"
        return f"❌ {err_msg}"

    # Dedup by date+source
    seen: set = set()
    deduped: List[DailyOTRecord] = []
    for r in all_records:
        key = (r.ad_date, r.source)
        if key not in seen:
            seen.add(key)
            deduped.append(r)
    deduped.sort(key=lambda r: r.ad_date)

    # Group by year × day_type
    year_day_map: Dict[int, Dict[str, List[DailyOTRecord]]] = {}
    for r in deduped:
        yr = r.ad_date.year
        year_day_map.setdefault(yr, {}).setdefault(r.day_type, []).append(r)

    lines: List[str] = ["📊 **加班費試算總表**\n"]
    grand_total = 0.0
    grand_ot_hours = 0.0

    # Self-validation accumulators
    val_total_by_type: Dict[str, float] = {}

    for yr in sorted(year_day_map.keys()):
        wage_raw = (monthly_wage_by_year or {}).get(yr, monthly_wage)
        wage_comp = _components_from_wage(wage_raw)
        hourly = wage_comp.total / 30.0 / 8.0
        yr_total = 0.0
        yr_ot_h = 0.0
        yr_roc = yr - 1911
        lines.append(f"▶ **{yr_roc}年度**")
        lines.append(f"  經常性薪資：{wage_comp.breakdown_str()}")
        lines.append(f"  時薪：{hourly:.4f} 元")

        for dtype in ["平日", "休息日", "例假日", "國定假日", "停班停課"]:
            recs = year_day_map[yr].get(dtype, [])
            if not recs:
                continue
            dt_min = sum(r.total_ot_min for r in recs)
            dt_hours = dt_min / 60.0
            dt_pay = sum(_calc_ot_pay_for_record(r, wage_comp) for r in recs)
            yr_total += dt_pay
            yr_ot_h += dt_hours
            val_total_by_type[dtype] = val_total_by_type.get(dtype, 0.0) + dt_pay
            lines.append(
                f"  {dtype}：{len(recs)} 天 / {dt_hours:.2f} 小時 → 加給 **{dt_pay:,.0f} 元**"
            )

        grand_total += yr_total
        grand_ot_hours += yr_ot_h
        lines.append(f"  小計：{yr_roc}年度加班加給 **{yr_total:,.0f} 元**（{yr_ot_h:.2f}h）\n")

    lines.append(f"─────────────────────────────")
    lines.append(f"**加班加給總計：{grand_total:,.0f} 元**（總計 {grand_ot_hours:.2f} 小時）")

    # ── 自我驗證 ──
    lines.append("\n**自我驗證**")
    val_ok = True
    for dtype, total_pay in val_total_by_type.items():
        # Recompute independently
        recs_all = [r for r in deduped if r.day_type == dtype]
        check_pay = 0.0
        for r in recs_all:
            w_raw = (monthly_wage_by_year or {}).get(r.ad_date.year, monthly_wage)
            check_pay += _calc_ot_pay_for_record(r, _components_from_wage(w_raw))
        diff = abs(total_pay - check_pay)
        ok = diff < 1.0
        val_ok = val_ok and ok
        lines.append(f"  {dtype}：主算 {total_pay:,.0f} / 複算 {check_pay:,.0f} → {'✅' if ok else '⚠️ 差異 ' + str(round(diff, 2))}")

    lines.append(f"\n整體驗算：{'✅ 一致' if val_ok else '⚠️ 有差異，請人工複核'}")
    lines.append(f"法條依據：勞基法第 24、39 條；停班停課雙倍依行政院人事行政總處規定")

    if errors:
        lines.append("\n⚠️ 部分檔案無法讀取：\n" + "\n".join(errors))

    return "\n".join(lines)


# ─── 自然語言解析 ─────────────────────────────────────────────────────────────

def _parse_task(task: str) -> Dict[str, Any]:
    """
    從自然語言任務文字解析計算參數。
    支援格式如：
      "月薪50000，休息日加班3小時"
      "到職日2020-03-01，計算特休"
      "月薪45000，資遣費，到職2018-01-01，離職2026-03-07"
      "加班費 月薪60000 例假日 8小時"
      "https://docs.google.com/spreadsheets/..."
    """
    params: Dict[str, Any] = {}

    # 本地檔案路徑（絕對路徑或含 .xlsx/.pdf 的相對路徑）
    file_paths = re.findall(r"(?:/[^\s,，；;]+\.(?:xlsx|xls|pdf)|[A-Za-z]:[^\s,，；;]+\.(?:xlsx|xls|pdf))", task, re.IGNORECASE)
    if file_paths:
        params["file_paths"] = file_paths
        params.setdefault("mode", "calc_file")

    # 案件目錄（整個資料夾）
    if any(k in task for k in ["案件目錄", "整個資料夾", "全部excel", "全部xlsx", "黃語玲案", "全部計算"]):
        params.setdefault("mode", "calc_dir")

    # Google Sheets URL
    if "docs.google.com" in task or "sheets.google" in task:
        url_m = re.search(r"https?://[^\s]+", task)
        if url_m:
            params["sheet_url"] = url_m.group(0)
            return params

    # 月薪（本薪）
    wage_m = re.search(r"(?:月薪|本薪|底薪|薪資|薪水)[^\d]*(\d[\d,]+)", task)
    if wage_m:
        params["monthly_wage"] = float(wage_m.group(1).replace(",", ""))

    # 伙食津貼 / 午餐加給
    meal_m = re.search(r"(?:伙食|午餐|餐飲|膳食)(?:津貼|加給|補助)?[^\d]*(\d[\d,]+)", task)
    if meal_m:
        params["meal_allowance"] = float(meal_m.group(1).replace(",", ""))

    # 全勤獎金
    att_m = re.search(r"(?:全勤)[^\d]*(\d[\d,]+)", task)
    if att_m:
        params["attendance_bonus"] = float(att_m.group(1).replace(",", ""))

    # 交通津貼
    trans_m = re.search(r"(?:交通)[^\d]*(\d[\d,]+)", task)
    if trans_m:
        params["transport_allowance"] = float(trans_m.group(1).replace(",", ""))

    # 職務加給 / 技術加給
    duty_m = re.search(r"(?:職務|技術)(?:加給|津貼)[^\d]*(\d[\d,]+)", task)
    if duty_m:
        params["duty_allowance"] = float(duty_m.group(1).replace(",", ""))

    # 其他固定津貼
    other_m = re.search(r"(?:其他(?:固定)?津貼|固定津貼)[^\d]*(\d[\d,]+)", task)
    if other_m:
        params["other_fixed_allowance"] = float(other_m.group(1).replace(",", ""))

    # 時薪（直接指定）
    hourly_m = re.search(r"(?:時薪)[^\d]*(\d[\d.]+)", task)
    if hourly_m:
        params["hourly_wage"] = float(hourly_m.group(1))

    # 加班時數
    hours_m = re.search(r"(\d+(?:\.\d+)?)\s*(?:小時|h|hr|hours?)", task)
    if hours_m:
        params["hours"] = float(hours_m.group(1))

    # 假別
    for kw, val in [("例假日", "例假日"), ("國定假日", "國定假日"), ("紀念日", "國定假日"),
                    ("休息日", "休息日"), ("平日", "平日")]:
        if kw in task:
            params["day_type"] = val
            break

    # 到職日
    hire_m = re.search(r"(?:到職|入職|到職日|受僱)[^\d]*(\d{4}[-/. ]\d{1,2}[-/. ]\d{1,2})", task)
    if hire_m:
        params["hire_date"] = _parse_date(hire_m.group(1))

    # 離職日
    term_m = re.search(r"(?:離職|終止|資遣日|結算)[^\d]*(\d{4}[-/. ]\d{1,2}[-/. ]\d{1,2})", task)
    if term_m:
        params["terminate_date"] = _parse_date(term_m.group(1))

    # 計算日期
    date_m = re.search(r"(?:計算日|參考日)[^\d]*(\d{4}[-/. ]\d{1,2}[-/. ]\d{1,2})", task)
    if date_m:
        params["ref_date"] = _parse_date(date_m.group(1))

    # 制度選擇
    if "舊制" in task:
        params["pension_choice"] = "old"
    elif "新制" in task:
        params["pension_choice"] = "new"

    # 功能判斷
    if any(k in task for k in ["資遣費", "資遣"]):
        params["mode"] = "severance"
    elif any(k in task for k in ["特休", "年假", "特別休假"]):
        params["mode"] = "annual_leave"
    elif any(k in task for k in ["加班費", "加班", "加給", "overtime"]):
        params["mode"] = "overtime"
    elif any(k in task for k in ["docs.google.com", "sheets"]):
        params["mode"] = "sheet"
    else:
        # 根據已解析欄位推斷
        if "hours" in params or "day_type" in params:
            params["mode"] = "overtime"
        elif "hire_date" in params and "monthly_wage" not in params:
            params["mode"] = "annual_leave"
        elif "hire_date" in params and "monthly_wage" in params:
            params.setdefault("mode", "severance")

    return params


# ─── 主要入口 ─────────────────────────────────────────────────────────────────

def _build_wage_components(params: Dict[str, Any]) -> Optional[WageComponents]:
    """從 params dict 組建 WageComponents，若無任何組成則回傳 None。"""
    base = params.get("monthly_wage")
    if base is None:
        return None
    return WageComponents(
        base=float(base),
        meal=float(params.get("meal_allowance", 0) or 0),
        attendance_bonus=float(params.get("attendance_bonus", 0) or 0),
        transport=float(params.get("transport_allowance", 0) or 0),
        duty_allowance=float(params.get("duty_allowance", 0) or 0),
        other_fixed=float(params.get("other_fixed_allowance", 0) or 0),
    )


def run(task: str, **kwargs) -> str:
    """主要入口：解析 task 文字，選擇計算模式並回傳結果。"""
    params = _parse_task(task)
    params.update({k: v for k, v in kwargs.items() if v is not None})

    mode = params.get("mode", "")
    ref_date = params.get("ref_date") or date.today()

    # Build WageComponents from all wage-related params
    wage_comp = _build_wage_components(params)

    # 本地檔案計算（xlsx + pdf）
    if mode == "calc_file" or "file_paths" in params:
        files = params.get("file_paths") or kwargs.get("file_paths") or []
        if not files:
            return "請提供本地檔案路徑（.xlsx 或 .pdf）"
        if not wage_comp:
            return "請提供月薪，例如：月薪 45000，伙食津貼 2400，檔案 /path/to/出勤.xlsx"
        wage_by_year = kwargs.get("monthly_wage_by_year")
        return calc_case_overtime(files, wage_comp, wage_by_year)

    # 案件資料夾（自動尋找所有 xlsx+pdf）
    if mode == "calc_dir":
        case_dir = params.get("case_dir") or kwargs.get("case_dir")
        if not case_dir:
            return "請提供案件資料夾路徑"
        all_files = []
        for p in Path(case_dir).rglob("*"):
            if p.suffix.lower() in (".xlsx", ".xls", ".pdf") and not p.name.startswith("."):
                all_files.append(str(p))
        if not all_files:
            return f"資料夾內找不到 xlsx/pdf 檔案：{case_dir}"
        if not wage_comp:
            return "請提供月薪"
        wage_by_year = kwargs.get("monthly_wage_by_year")
        return calc_case_overtime(all_files, wage_comp, wage_by_year)

    # Google Sheets
    if "sheet_url" in params or mode == "sheet":
        url = params.get("sheet_url", "")
        if not url:
            url_m = re.search(r"https?://[^\s]+", task)
            url = url_m.group(0) if url_m else ""
        if not url:
            return "請提供 Google Sheets 的公開分享連結。"
        return calc_from_sheet(url, ref_date)

    # 加班費
    if mode == "overtime":
        if not wage_comp:
            if "hourly_wage" in params:
                # 從時薪反推月薪（月薪 = 時薪 × 240）
                wage_comp = WageComponents(base=params["hourly_wage"] * 240)
            else:
                return "請提供月薪金額，例如：月薪 50000，伙食津貼 2400，休息日加班 3 小時"
        hours = params.get("hours", 2.0)
        day_type = params.get("day_type", "平日")
        r = calc_overtime(wage_comp, hours, day_type, ref_date)
        return _fmt_overtime(r)

    # 特休假
    if mode == "annual_leave":
        hire_date = params.get("hire_date")
        if not hire_date:
            return "請提供到職日，例如：到職日 2020-03-01，計算特休"
        r = calc_annual_leave(hire_date, ref_date)
        return _fmt_annual_leave(r)

    # 資遣費
    if mode == "severance":
        hire_date = params.get("hire_date")
        if not wage_comp or not hire_date:
            return "請提供月薪（平均工資）與到職日，例如：月薪 50000，到職 2015-01-01，資遣費"
        terminate_date = params.get("terminate_date") or ref_date
        pension_choice = params.get("pension_choice", "auto")
        # 資遣費平均工資 = 經常性薪資合計
        r = calc_severance(wage_comp.total, hire_date, terminate_date, pension_choice)
        note = f"（平均工資計算基礎：{wage_comp.breakdown_str()}）\n"
        return note + _fmt_severance(r)

    # 兜底：顯示說明
    return (
        "📋 **勞動基準法計算說明**\n\n"
        "支援以下計算，請以自然語言描述：\n\n"
        "**加班費**：月薪 50000，伙食津貼 2400，休息日加班 3 小時\n"
        "**特休假**：到職日 2020-03-01，計算特休\n"
        "**資遣費**：月薪 45000，伙食津貼 2400，全勤 500，到職 2018-01-01，離職 2026-03-07，資遣費\n"
        "**試算表代算**：貼上 Google Sheets 公開分享連結\n\n"
        "假別選項：平日 / 休息日 / 例假日 / 國定假日\n"
        "資遣費制度：可加「舊制」或「新制」指定，否則自動判斷混合制\n\n"
        "**📌 經常性薪資說明（勞基法第 2 條第 3 款）**\n"
        "✅ 算入工資：本薪、伙食津貼（每月固定）、全勤獎金（每月固定）、\n"
        "           職務加給、技術加給、交通津貼（每月固定）、其他固定津貼\n"
        "❌ 不算入：年終獎金、三節獎金、績效獎金（非固定）、差旅費、健保勞保費\n"
        "（依勞基法施行細則第 10 條）"
    )


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="台灣勞動基準法計算器")
    ap.add_argument("--task", default="", help="計算指令（自然語言）")
    ap.add_argument("--mode", choices=["overtime", "annual_leave", "severance", "sheet"], help="強制指定計算模式")
    ap.add_argument("--monthly-wage", type=float, help="本薪/底薪（元）")
    ap.add_argument("--meal-allowance", type=float, default=0, help="伙食津貼/午餐加給（每月固定給付，算入工資）")
    ap.add_argument("--attendance-bonus", type=float, default=0, help="全勤獎金（每月固定）")
    ap.add_argument("--transport-allowance", type=float, default=0, help="交通津貼（每月固定）")
    ap.add_argument("--duty-allowance", type=float, default=0, help="職務加給/技術加給")
    ap.add_argument("--other-fixed-allowance", type=float, default=0, help="其他每月固定津貼")
    ap.add_argument("--day-type", help="假別：平日/休息日/例假日/國定假日")
    ap.add_argument("--hours", type=float, help="加班時數")
    ap.add_argument("--hire-date", help="到職日（YYYY-MM-DD）")
    ap.add_argument("--terminate-date", help="離職/資遣日（YYYY-MM-DD）")
    ap.add_argument("--ref-date", help="計算參考日（YYYY-MM-DD，預設今日）")
    ap.add_argument("--pension-choice", choices=["old", "new", "auto"], default="auto", help="退休金制度")
    ap.add_argument("--sheet-url", help="Google Sheets 試算表 URL")
    ap.add_argument("--file", nargs="+", help="本地 xlsx/pdf 檔案路徑（可多個）")
    ap.add_argument("--case-dir", help="案件資料夾路徑（自動讀取所有 xlsx+pdf）")
    ap.add_argument("--wage-by-year", help='各年度月薪 JSON，例如 {"2021":45000,"2022":47000}')
    ap.add_argument("--json", action="store_true", help="輸出 JSON 格式")
    args = ap.parse_args()

    kwargs = {}
    if args.mode:
        kwargs["mode"] = args.mode
    if args.monthly_wage:
        kwargs["monthly_wage"] = args.monthly_wage
    if args.meal_allowance:
        kwargs["meal_allowance"] = args.meal_allowance
    if args.attendance_bonus:
        kwargs["attendance_bonus"] = args.attendance_bonus
    if args.transport_allowance:
        kwargs["transport_allowance"] = args.transport_allowance
    if args.duty_allowance:
        kwargs["duty_allowance"] = args.duty_allowance
    if args.other_fixed_allowance:
        kwargs["other_fixed_allowance"] = args.other_fixed_allowance
    if args.day_type:
        kwargs["day_type"] = args.day_type
    if args.hours:
        kwargs["hours"] = args.hours
    if args.hire_date:
        kwargs["hire_date"] = _parse_date(args.hire_date)
    if args.terminate_date:
        kwargs["terminate_date"] = _parse_date(args.terminate_date)
    if args.ref_date:
        kwargs["ref_date"] = _parse_date(args.ref_date)
    if args.pension_choice != "auto":
        kwargs["pension_choice"] = args.pension_choice
    if args.sheet_url:
        kwargs["sheet_url"] = args.sheet_url
    if args.file:
        kwargs["file_paths"] = args.file
        kwargs.setdefault("mode", "calc_file")
    if args.case_dir:
        kwargs["case_dir"] = args.case_dir
        kwargs.setdefault("mode", "calc_dir")
    if args.wage_by_year:
        try:
            raw = json.loads(args.wage_by_year)
            kwargs["monthly_wage_by_year"] = {int(k): float(v) for k, v in raw.items()}
        except Exception:
            print(f"⚠️ --wage-by-year 格式錯誤，忽略", file=sys.stderr)

    task = args.task or ""
    result = run(task, **kwargs)

    if args.json:
        print(json.dumps({"result": result}, ensure_ascii=False, indent=2))
    else:
        print(result)


if __name__ == "__main__":
    main()
