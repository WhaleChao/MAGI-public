#!/usr/bin/env python3
"""Classify Supreme Court judgments that mention interpreters.

The classifier is deliberately evidence-led: every row keeps short snippets
around "通譯" so the resulting table can be reviewed and corrected.
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


DEFAULT_INPUT_DIR = Path("/Users/ai/Desktop/最高法院_通譯_TXT")
DEFAULT_INPUT_CANDIDATES = (
    DEFAULT_INPUT_DIR / "完整812" / "TXT",
    DEFAULT_INPUT_DIR / "TXT",
    DEFAULT_INPUT_DIR,
    Path("/Users/ai/Desktop/AGENT TEST DATA/最高法院_通譯_TXT/完整812/TXT"),
)


QUALITY_PATTERNS = [
    r"通譯[^。；，\n]{0,80}(不實|虛偽|錯誤|誤譯|漏譯|未如實|不正確|不公|偏頗|不符|資格|適任)",
    r"(翻譯|傳譯|譯文)[^。；，\n]{0,80}(不實|虛偽|錯誤|誤譯|漏譯|未如實|不正確|不公|偏頗|不符)",
    r"(不實|虛偽|錯誤|誤譯|漏譯|未如實|不正確|不公|偏頗|不符)[^。；，\n]{0,80}(通譯|翻譯|傳譯|譯文)",
]

NO_INTERPRETER_PATTERNS = [
    r"(無|未|沒有|未經|未有|不曾|未命|未置|未由)[^。；，\n]{0,30}通譯",
    r"通譯[^。；，\n]{0,40}(未在場|未到場|未傳譯|未翻譯|未協助)",
    r"(語言不通|不通曉|不諳|聽不懂|無法理解|不能理解)[^。；，\n]{0,80}(通譯|訊問|審理|警詢|偵訊)",
]

EVIDENCE_TRANSLATION_PATTERNS = [
    r"(經|由|請|囑請|委由|交由)[^。；，\n]{0,30}通譯[^。；，\n]{0,60}(翻譯|傳譯|辨識|譯出|回復|譯文)",
    r"通譯[^。；，\n]{0,60}(當庭翻譯|翻譯結果|傳譯|譯為|辨識|回復內容)",
    r"(錄音|錄影|勘驗|對話|LINE|光碟)[^。；，\n]{0,120}(通譯|翻譯|譯文|傳譯)",
]

LEGAL_TEMPLATE_PATTERNS = [
    r"證人、鑑定人、通譯",
    r"證言、鑑定或通譯",
    r"法官、書記官、通譯",
    r"通譯及其他非當事人",
    r"刑事訴訟法第?420條[^。；\n]{0,120}通譯",
    r"第420條[^。；\n]{0,120}通譯",
    r"第403條[^。；\n]{0,120}通譯",
    r"第415條[^。；\n]{0,120}通譯",
]

TRIAL_ACCESS_PATTERNS = [
    r"通譯[^。；，\n]{0,80}(電腦螢幕|螢幕|筆錄|辯護|防禦權|訴訟防禦|權利)",
    r"(電腦螢幕|螢幕|筆錄|辯護|防禦權|訴訟防禦)[^。；，\n]{0,80}通譯",
]

REJECTED_PATTERNS = [
    r"(不足採|無理由|駁回|誤會|難認|尚非|不得指為違法|不能因此指為違法|無違法|無不合|不合法|不符合|無從動搖|並無理由)",
]

ACCEPTED_PATTERNS = [
    r"(有理由|撤銷發回|撤銷，發回|違法|調查未盡|理由不備|影響於判決|非無疑義|應予究明|未洽)",
]

FACTUAL_INTERPRETER_MARKERS = [
    r"通譯(即|並未|未|不|錯|漏|在場|到場|翻譯|傳譯|譯文|陳述|證人)",
    r"(翻譯|傳譯|譯文|語言不通|不通曉|不諳|外語|越南語|印尼語|泰語|英文)",
    r"(當庭|警詢|偵訊|訊問|審理|勘驗|錄音|錄影|LINE)[^。；\n]{0,120}(通譯|翻譯|傳譯|譯文)",
]


@dataclass
class ClassifiedCase:
    txt_index: str
    authoritative_index: str
    court_no: str
    date: str
    cause: str
    outcome: str
    primary_category: str
    categories: str
    issue_role: str
    issue_result: str
    interpreter_marker: str
    confidence: str
    prior_case_no: str
    snippets: str
    source_file: str
    pdf_file: str


def compact(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").replace("\u3000", " ")).strip()


def compile_any(patterns: Iterable[str], text: str) -> bool:
    return any(re.search(pattern, text) for pattern in patterns)


def is_pure_legal_template_context(context: str) -> bool:
    """Return True for generic legal formula mentions, not case-specific issues."""
    if not compile_any(LEGAL_TEMPLATE_PATTERNS, context):
        return False
    return not compile_any(FACTUAL_INTERPRETER_MARKERS, context)


def first_match(pattern: str, text: str) -> str:
    match = re.search(pattern, text, flags=re.S)
    return compact(match.group(1)) if match else ""


def extract_metadata(text: str, fallback_name: str = "") -> dict:
    return {
        "court_no": first_match(r"裁判字號：\s*([^\n]+)", text) or fallback_name,
        "date": first_match(r"裁判日期：\s*([^\n]+)", text),
        "cause": first_match(r"裁判案由：\s*([^\n]+)", text),
    }


def extract_main_text(text: str) -> str:
    match = re.search(r"\n主文\s*\n(?P<main>.*?)(?:\n理由|\n事實及理由|\n事實)", text, flags=re.S)
    if not match:
        return ""
    return compact(match.group("main"))


def classify_outcome(main_text: str) -> str:
    main = compact(main_text)
    if not main:
        return "未抽出主文"
    has_revoke = "撤銷" in main
    has_remand = "發回" in main
    has_dismiss = "駁回" in main
    if has_revoke and has_remand and has_dismiss:
        return "部分撤銷發回、部分駁回"
    if has_revoke and has_remand:
        return "撤銷發回"
    if "原判決撤銷" in main:
        return "原判決撤銷"
    if "再抗告駁回" in main:
        return "再抗告駁回"
    if "抗告駁回" in main:
        return "抗告駁回"
    if "上訴駁回" in main:
        return "上訴駁回"
    if "聲請駁回" in main:
        return "聲請駁回"
    if has_dismiss:
        return "駁回"
    return main[:80]


def extract_prior_case_no(text: str) -> str:
    intro = compact(text[:1800])
    matches = re.findall(r"[（(]([^（）()]{0,12}\d{2,3}\s*年度[^（）()]{0,45}?字第\s*\d+\s*號[^（）()]*)[）)]", intro)
    cleaned = []
    for match in matches:
        item = compact(match)
        if "年度" in item and "字第" in item and item not in cleaned:
            cleaned.append(item)
    return "；".join(cleaned[:3])


def find_contexts(text: str, keyword: str = "通譯", window: int = 180, max_items: int = 30) -> list[str]:
    contexts: list[str] = []
    seen: set[str] = set()
    for match in re.finditer(re.escape(keyword), text):
        start = max(0, match.start() - window)
        end = min(len(text), match.end() + window)
        snippet = compact(text[start:end])
        snippet = snippet.strip("，。；、 ")
        key = snippet[:80]
        if key in seen:
            continue
        seen.add(key)
        contexts.append(snippet)
        if len(contexts) >= max_items:
            break
    return contexts


def classify_contexts(contexts: list[str], outcome: str) -> tuple[str, str, str, str, str, str]:
    if not contexts:
        return "缺全文", "缺全文", "非通譯爭點", "非通譯爭點", "缺全文", "低"
    joined = "\n".join(contexts)
    material_contexts = [c for c in contexts if not is_pure_legal_template_context(c)]
    material_joined = "\n".join(material_contexts)
    has_quality = compile_any(QUALITY_PATTERNS, material_joined)
    has_no_interpreter = compile_any(NO_INTERPRETER_PATTERNS, material_joined)
    has_trial_access = compile_any(TRIAL_ACCESS_PATTERNS, material_joined)
    has_evidence = compile_any(EVIDENCE_TRANSLATION_PATTERNS, material_joined)
    has_legal = compile_any(LEGAL_TEMPLATE_PATTERNS, joined)

    material: list[str] = []
    if has_quality:
        material.append("通譯/翻譯品質或真實性爭議")
    if has_no_interpreter:
        material.append("未使用或未充分使用通譯爭議")
    if has_trial_access:
        material.append("通譯參與之法庭程序/辯護權爭議")
    if has_evidence:
        material.append("外語證據翻譯/勘驗譯文")
    if has_legal:
        material.append("法條或程序清單引用")
    if not material:
        material.append("單純提及或其他")

    issue_categories = [c for c in material if c not in {"外語證據翻譯/勘驗譯文", "法條或程序清單引用"}]
    is_interpreter_issue = any(c in material for c in [
        "通譯/翻譯品質或真實性爭議",
        "未使用或未充分使用通譯爭議",
        "通譯參與之法庭程序/辯護權爭議",
    ])
    role = "通譯為上訴/抗告/再審爭點" if is_interpreter_issue else "非通譯爭點"
    if is_interpreter_issue:
        interpreter_marker = "實質通譯爭點"
    elif has_legal:
        interpreter_marker = "僅條文引用"
    else:
        interpreter_marker = "含通譯文字"

    rejected = compile_any(REJECTED_PATTERNS, material_joined)
    accepted = compile_any(ACCEPTED_PATTERNS, material_joined) or "撤銷發回" in outcome
    if not is_interpreter_issue:
        issue_result = "非通譯爭點"
        confidence = "中" if has_legal else "中低"
    elif accepted and "撤銷發回" in outcome and not rejected:
        issue_result = "疑似採納或與撤銷發回相關"
        confidence = "中"
    elif rejected or "駁回" in outcome:
        issue_result = "未採或駁回"
        confidence = "中高"
    else:
        issue_result = "需人工複核"
        confidence = "中"

    primary = issue_categories[0] if issue_categories else material[0]
    return primary, "；".join(material), role, issue_result, interpreter_marker, confidence


def classify_file(path: Path) -> ClassifiedCase:
    text = path.read_text(encoding="utf-8", errors="replace")
    meta = extract_metadata(text, fallback_name=path.stem)
    main_text = extract_main_text(text)
    outcome = classify_outcome(main_text)
    contexts = find_contexts(text)
    primary, categories, role, issue_result, interpreter_marker, confidence = classify_contexts(contexts, outcome)
    txt_index, authoritative_index = source_indexes(path)
    return ClassifiedCase(
        txt_index=txt_index,
        authoritative_index=authoritative_index,
        court_no=meta["court_no"],
        date=meta["date"],
        cause=meta["cause"],
        outcome=outcome,
        primary_category=primary,
        categories=categories,
        issue_role=role,
        issue_result=issue_result,
        interpreter_marker=interpreter_marker,
        confidence=confidence,
        prior_case_no=extract_prior_case_no(text),
        snippets=format_snippets(representative_contexts(contexts)),
        source_file=str(path),
        pdf_file=str(resolve_pdf_path(path)),
    )


def resolve_pdf_path(path: Path) -> Path | str:
    pdf_name = path.with_suffix(".pdf").name
    for candidate in (
        path.parent / "PDF" / pdf_name,
        path.parent.parent / "PDF" / pdf_name,
        path.with_suffix(".pdf"),
    ):
        if candidate.exists():
            return candidate
    return ""


def source_indexes(path: Path) -> tuple[str, str]:
    txt_match = re.match(r"^(\d{4})_", path.name)
    txt_index = txt_match.group(1) if txt_match else ""
    mapping_path = path.parent / "重新編號對照表.csv"
    if not txt_index or not mapping_path.exists():
        return txt_index, ""
    try:
        with mapping_path.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                if row.get("new_index", "").zfill(4) == txt_index:
                    return txt_index, str(row.get("old_authoritative_index", "")).zfill(4)
    except Exception:
        return txt_index, ""
    return txt_index, ""


def representative_contexts(contexts: list[str], limit: int = 3) -> list[str]:
    material = [c for c in contexts if not is_pure_legal_template_context(c)]
    legal = [c for c in contexts if is_pure_legal_template_context(c)]
    selected = (material + legal)[:limit]
    return selected


def format_snippets(contexts: list[str]) -> str:
    if not contexts:
        return ""
    parts = []
    for idx, context in enumerate(contexts, start=1):
        context = context.replace("通譯", "【通譯】")
        parts.append(f"【原文摘錄 {idx}】{context}")
    return "\n\n".join(parts)


def dedupe_cases(rows: list[ClassifiedCase]) -> list[ClassifiedCase]:
    by_key: dict[str, ClassifiedCase] = {}
    for row in rows:
        key = row.court_no or row.source_file
        old = by_key.get(key)
        if not old or len(row.snippets) > len(old.snippets):
            by_key[key] = row
    return list(by_key.values())


def write_csv(rows: list[ClassifiedCase], path: Path) -> None:
    fields = list(ClassifiedCase.__dataclass_fields__.keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.__dict__)


def write_markdown(rows: list[ClassifiedCase], path: Path) -> None:
    headers = ["最高法院裁判字號", "裁判日期", "案由", "主分類", "法院結果", "通譯爭點處理", "摘錄"]
    lines = [
        "# 最高法院「通譯」裁判分類表",
        "",
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        values = [
            row.court_no,
            row.date,
            row.cause,
            row.primary_category,
            row.outcome,
            row.issue_result,
            row.snippets[:220].replace("\n", "<br>"),
        ]
        escaped = [compact(v).replace("|", "｜") for v in values]
        lines.append("| " + " | ".join(escaped) + " |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_xlsx(rows: list[ClassifiedCase], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    fields = list(ClassifiedCase.__dataclass_fields__.keys())
    labels = {
        "txt_index": "目前TXT編號",
        "authoritative_index": "原812清單序號",
        "court_no": "最高法院裁判字號",
        "date": "裁判日期",
        "cause": "案由",
        "outcome": "法院判決/裁定結果",
        "primary_category": "主分類",
        "categories": "全部分類",
        "issue_role": "通譯角色",
        "issue_result": "通譯爭點處理",
        "interpreter_marker": "通譯判讀標記",
        "confidence": "分類信心",
        "prior_case_no": "前審/相關案號",
        "snippets": "通譯相關原文摘錄",
        "source_file": "來源檔案",
        "pdf_file": "PDF檔案",
    }
    category_fills = {
        "通譯/翻譯品質或真實性爭議": "FCE4D6",
        "未使用或未充分使用通譯爭議": "FFF2CC",
        "通譯參與之法庭程序/辯護權爭議": "E4DFEC",
        "外語證據翻譯/勘驗譯文": "DDEBF7",
        "法條或程序清單引用": "E2F0D9",
        "單純提及或其他": "F2F2F2",
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "通譯分類"
    ws.append([labels[f] for f in fields])
    for row in rows:
        ws.append([getattr(row, f) for f in fields])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {
        "A": 14, "B": 16, "C": 34, "D": 18, "E": 24, "F": 22,
        "G": 28, "H": 46, "I": 24, "J": 24, "K": 12, "L": 44,
        "M": 24, "N": 100, "O": 60, "P": 60,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        category = str(ws.cell(row=idx, column=7).value or "")
        fill = PatternFill("solid", fgColor=category_fills.get(category, "FFFFFF"))
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[idx].height = 96
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("分類統計")
    summary.append(["主分類", "件數"])
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.primary_category] = counts.get(row.primary_category, 0) + 1
    for key, value in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
        summary.append([key, value])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    summary.column_dimensions["A"].width = 40
    summary.column_dimensions["B"].width = 12

    chart = BarChart()
    chart.title = "最高法院提及通譯裁判分類"
    chart.y_axis.title = "分類"
    chart.x_axis.title = "件數"
    data = Reference(summary, min_col=2, min_row=1, max_row=summary.max_row)
    cats = Reference(summary, min_col=1, min_row=2, max_row=summary.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    summary.add_chart(chart, "D2")

    legend = wb.create_sheet("分類說明")
    legend.append(["分類", "判斷方式"])
    legend_rows = [
        ("通譯/翻譯品質或真實性爭議", "原文提到通譯、翻譯、傳譯、譯文有不實、虛偽、錯誤、誤譯、漏譯、未如實、不公、偏頗等爭議。"),
        ("未使用或未充分使用通譯爭議", "原文提到無通譯、未經通譯、未命通譯、語言不通卻未協助等程序爭議。"),
        ("通譯參與之法庭程序/辯護權爭議", "原文提到通譯參與造成筆錄確認、電腦螢幕、辯護權或防禦權問題。"),
        ("外語證據翻譯/勘驗譯文", "原文提到錄音、錄影、LINE、勘驗等外語內容經通譯翻譯或辨識，但未必主張違法。"),
        ("法條或程序清單引用", "原文只是引用刑事訴訟法或程序清單，如證人、鑑定人、通譯等，通常不是個案通譯爭點。"),
        ("單純提及或其他", "含通譯字樣，但不符合前述類型，需要個別閱讀原文摘錄。"),
    ]
    for item in legend_rows:
        legend.append(list(item))
    for cell in legend[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    legend.column_dimensions["A"].width = 36
    legend.column_dimensions["B"].width = 100
    for row in legend.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def resolve_input_dir(raw: str = "") -> Path:
    if raw:
        return Path(raw).expanduser()
    for candidate in DEFAULT_INPUT_CANDIDATES:
        if candidate.exists() and any(candidate.glob("*.txt")):
            return candidate
    return DEFAULT_INPUT_DIR


def main() -> int:
    parser = argparse.ArgumentParser(description="分類最高法院裁判中提到通譯的段落")
    parser.add_argument("--input-dir", default="")
    parser.add_argument("--output-prefix", default="")
    args = parser.parse_args()

    input_dir = resolve_input_dir(args.input_dir)
    prefix = Path(args.output_prefix) if args.output_prefix else input_dir / "最高法院_通譯_分類表"
    rows = [classify_file(path) for path in sorted(input_dir.glob("*.txt"))]
    rows = dedupe_cases(rows)
    rows.sort(key=lambda row: (row.date, row.court_no), reverse=True)

    write_csv(rows, prefix.with_suffix(".csv"))
    write_markdown(rows, prefix.with_suffix(".md"))
    write_xlsx(rows, prefix.with_suffix(".xlsx"))
    print(f"classified={len(rows)}")
    print(f"csv={prefix.with_suffix('.csv')}")
    print(f"xlsx={prefix.with_suffix('.xlsx')}")
    print(f"md={prefix.with_suffix('.md')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
