from datetime import date


def test_period_for_settlement_month_uses_26_to_25_window():
    from api.osc.accounting_bonus import default_settlement_month, period_for_settlement_month

    start, end, key = period_for_settlement_month("2026-05")
    assert key == "2026-05"
    assert start.isoformat() == "2026-04-26"
    assert end.isoformat() == "2026-05-25"
    assert default_settlement_month(date(2026, 5, 24), catch_up=True) == "2026-05"
    assert default_settlement_month(date(2026, 6, 3), catch_up=True) == "2026-05"
    assert default_settlement_month(date(2026, 6, 23), catch_up=True) == "2026-05"
    assert default_settlement_month(date(2026, 6, 24), catch_up=True) == "2026-06"
    assert default_settlement_month(date(2026, 5, 12), catch_up=True) == "2026-04"
    assert default_settlement_month(date(2026, 5, 12), catch_up=False) is None


def test_laf_fee_basis_defaults_to_current_settlement_period(monkeypatch):
    from api.osc.accounting_bonus import _laf_fee_basis_start

    monkeypatch.delenv("MAGI_ACCOUNTING_BONUS_ALLOW_LOOKBACK", raising=False)
    monkeypatch.delenv("MAGI_ACCOUNTING_BONUS_LAF_FEE_LOOKBACK_START", raising=False)
    assert _laf_fee_basis_start("2026-05", date(2026, 4, 26), date(2026, 5, 25)) == date(2026, 4, 26)


def test_accounting_transaction_write_requires_operator_when_login_enabled(monkeypatch):
    from flask import Flask
    from flask_login import LoginManager, UserMixin, login_user
    from api.blueprints import osc_accounting as mod

    app = Flask(__name__)
    app.config["TESTING"] = True
    app.config["LOGIN_DISABLED"] = False
    app.secret_key = "test"
    login = LoginManager()
    login.init_app(app)

    class Viewer(UserMixin):
        id = "viewer"
        role = "viewer"

    @login.user_loader
    def _load_user(_user_id):
        return Viewer()

    @app.route("/login")
    def _login():
        login_user(Viewer())
        return "ok"

    monkeypatch.setattr(
        mod,
        "_get_osc_helpers",
        lambda: (
            lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("DB helper should not be called")),
            str,
            lambda *args, **kwargs: None,
            lambda raw: raw,
            int,
        ),
    )
    app.register_blueprint(mod.osc_accounting_bp)
    client = app.test_client()
    client.get("/login")

    r = client.post("/api/osc/accounting/transactions", json={"amount": 100, "type": "收入"})

    assert r.status_code == 403
    assert r.get_json()["error"] == "forbidden"


def test_laf_fee_basis_lookback_is_explicit_opt_in(monkeypatch):
    from api.osc.accounting_bonus import _laf_fee_basis_start

    monkeypatch.setenv("MAGI_ACCOUNTING_BONUS_ALLOW_LOOKBACK", "1")
    monkeypatch.setenv("MAGI_ACCOUNTING_BONUS_LAF_FEE_LOOKBACK_START", "2026-01-01")
    assert _laf_fee_basis_start("2026-05", date(2026, 4, 26), date(2026, 5, 25)) == date(2026, 1, 1)


def test_query_laf_debt_fee_rows_filters_to_debt_laf_income(monkeypatch):
    from api.osc import accounting_bonus as mod

    candidate_rows = [
        {
            "id": 1,
            "date": "2026-05-20",
            "type": "收入",
            "category": "法扶案件",
            "description": "1150317-E-009 邱衣萱｜預付酬金",
            "amount": 4000,
            "case_type": "消費者債務清理",
            "case_category": "法律扶助案件",
            "case_reason": "更生",
            "case_number": "2026-0024",
            "client_name": "邱衣萱",
        },
        {
            "id": 2,
            "date": "2026-05-20",
            "type": "收入",
            "category": "法扶案件",
            "description": "1150320-E-014 劉信義｜預付酬金",
            "amount": 15000,
            "case_type": "刑事",
            "case_category": "法律扶助案件",
            "case_reason": "殺人",
        },
    ]

    def fake_exec(sql, params=(), fetch="none"):
        if "FROM case_transactions t" in sql:
            return candidate_rows, {}
        if "FROM cases" in sql:
            return None, {}
        raise AssertionError(sql)

    monkeypatch.setattr(mod, "_get_osc_helpers", lambda: fake_exec)
    rows = mod.query_laf_debt_fee_rows(date(2026, 5, 1), date(2026, 5, 31))
    assert len(rows) == 1
    assert rows[0]["amount"] == 4000
    assert rows[0]["client_name"] == "邱衣萱"


def test_query_laf_debt_fee_rows_matches_single_debt_case_by_client_name(monkeypatch):
    from api.osc import accounting_bonus as mod

    candidate_rows = [
        {
            "id": 8,
            "date": "2026-05-20",
            "type": "收入",
            "category": "法扶案件",
            "description": "1150304-E-003 黃鎮洲｜預付酬金",
            "amount": 4000,
            "case_type": None,
            "case_category": None,
            "case_reason": None,
            "case_number": None,
            "client_name": None,
            "legal_aid_number": None,
            "laf_case_no": None,
        }
    ]
    matching_case = {
        "id": "2026-0031",
        "case_number": "2026-0031",
        "client_name": "黃鎮洲",
        "case_type": "消費者債務清理",
        "case_category": "法律扶助案件",
        "case_reason": "更生",
        "folder_path": "",
        "folder_name": "",
        "legal_aid_number": "",
        "laf_case_no": "",
    }

    def fake_exec(sql, params=(), fetch="none"):
        if "FROM case_transactions t" in sql:
            return candidate_rows, {}
        if "WHERE legal_aid_number=%s" in sql:
            return None, {}
        if "WHERE client_name=%s" in sql:
            assert params == ("黃鎮洲",)
            return [matching_case], {}
        raise AssertionError(sql)

    monkeypatch.setattr(mod, "_get_osc_helpers", lambda: fake_exec)
    rows = mod.query_laf_debt_fee_rows(date(2026, 5, 1), date(2026, 5, 31))
    assert len(rows) == 1
    assert rows[0]["case_number"] == "2026-0031"
    assert rows[0]["client_name"] == "黃鎮洲"
    assert rows[0]["amount"] == 4000


def test_query_laf_debt_fee_rows_can_scope_by_import_source_month(monkeypatch):
    from api.osc import accounting_bonus as mod

    candidate_rows = [
        {
            "id": 18,
            "date": "2026-05-25",
            "type": "收入",
            "category": "法扶案件",
            "description": "1150303-I-004 林亮宏｜預付酬金",
            "amount": 4000,
            "case_type": "消費者債務清理",
            "case_category": "法律扶助案件",
            "case_reason": "清算",
            "case_number": "2026-0022",
            "client_name": "林亮宏",
        }
    ]

    def fake_exec(sql, params=(), fetch="none"):
        if "FROM case_transactions t" in sql:
            assert "accounting_import_records" in sql
            assert params[:4] == ("colleague_google_sheet", "2026-06", "2026-05-26", "2026-06-25")
            return candidate_rows, {}
        if "FROM cases" in sql:
            return None, {}
        raise AssertionError(sql)

    monkeypatch.setattr(mod, "_get_osc_helpers", lambda: fake_exec)
    rows = mod.query_laf_debt_fee_rows(
        date(2026, 5, 26),
        date(2026, 6, 25),
        source_month="2026-06",
    )
    assert len(rows) == 1
    assert rows[0]["date"] == "2026-05-25"


def test_query_totals_before_bonus_can_scope_by_import_source_month(monkeypatch):
    from api.osc import accounting_bonus as mod

    def fake_exec(sql, params=(), fetch="none"):
        assert "accounting_import_records" in sql
        assert "air.source_month=%s" in sql
        assert params[:4] == ("colleague_google_sheet", "2026-06", "2026-05-26", "2026-06-25")
        return {"income_total": 50000, "expense_total": 1290}, {}

    monkeypatch.setattr(mod, "_get_osc_helpers", lambda: fake_exec)
    totals = mod._query_totals_before_bonus(date(2026, 5, 26), date(2026, 6, 25), source_month="2026-06")
    assert totals == {"income_total": 50000, "expense_total": 1290}


def test_calculate_monthly_bonus_waits_when_no_laf_fee(monkeypatch):
    from api.osc import accounting_bonus as mod

    calls = []

    def fake_exec(sql, params=(), fetch="none"):
        calls.append((sql, params, fetch))
        if sql.strip().startswith("CREATE TABLE"):
            return {}, {}
        if sql.strip().startswith("SHOW COLUMNS"):
            return [{"Field": "xlsx_path"}], {}
        if "income_total" in sql and "expense_total" in sql:
            return {"income_total": 30000, "expense_total": 10000}, {}
        if "FROM case_transactions t" in sql:
            return [], {}
        if "SELECT id FROM case_transactions" in sql:
            return None, {}
        if "INSERT INTO accounting_monthly_bonus_runs" in sql:
            return {"rowcount": 1}, {}
        return {"rowcount": 0}, {}

    monkeypatch.setattr(mod, "_get_osc_helpers", lambda: fake_exec)
    monkeypatch.setattr(mod, "refresh_accounting_import_for_period", lambda *a, **k: [])
    result = mod.calculate_monthly_bonus(month="2026-05", commit=True, refresh_import=False)

    assert result["status"] == "waiting_laf_fee"
    assert result["legal_aid_bonus_amount"] == 0
    assert result["case_bonus_employee_amount"] == 0
    assert any("accounting_monthly_bonus_runs" in sql for sql, _, _ in calls)


def test_ensure_bonus_schema_adds_xlsx_path_for_existing_table(monkeypatch):
    from api.osc import accounting_bonus as mod

    calls = []

    def fake_exec(sql, params=(), fetch="none"):
        calls.append((sql, params, fetch))
        if sql.strip().startswith("SHOW COLUMNS"):
            return [{"Field": "month_key"}], {}
        return {"rowcount": 0}, {}

    monkeypatch.setattr(mod, "_get_osc_helpers", lambda: fake_exec)
    mod.ensure_bonus_schema()

    assert any("ADD COLUMN xlsx_path" in sql for sql, _, _ in calls)


def test_calculate_monthly_bonus_posts_two_expenses(monkeypatch):
    from api.osc import accounting_bonus as mod

    inserts = []

    def fake_exec(sql, params=(), fetch="none"):
        if sql.strip().startswith("CREATE TABLE"):
            return {}, {}
        if sql.strip().startswith("SHOW COLUMNS"):
            return [{"Field": "xlsx_path"}], {}
        if "income_total" in sql and "expense_total" in sql:
            return {"income_total": 30000, "expense_total": 10000}, {}
        if "FROM case_transactions t" in sql:
            return [
                {
                    "id": 9,
                    "date": "2026-05-20",
                    "type": "收入",
                    "category": "法扶案件",
                    "description": "1150317-E-009 邱衣萱｜預付酬金",
                    "amount": 10000,
                    "case_type": "消費者債務清理",
                    "case_category": "法律扶助案件",
                    "case_reason": "更生",
                    "case_number": "2026-0024",
                    "client_name": "邱衣萱",
                    "laf_case_no": "1150317-E-009",
                }
            ], {}
        if "SELECT id FROM case_transactions" in sql:
            return None, {}
        if "INSERT INTO case_transactions" in sql:
            inserts.append(params)
            return {"lastrowid": 100 + len(inserts)}, {}
        if "INSERT INTO accounting_monthly_bonus_runs" in sql:
            return {"rowcount": 1}, {}
        return {"rowcount": 0}, {}

    monkeypatch.setattr(mod, "_get_osc_helpers", lambda: fake_exec)
    monkeypatch.setattr(mod, "refresh_accounting_import_for_period", lambda *a, **k: [{"ok": True, "month": "2026-05"}])
    result = mod.calculate_monthly_bonus(month="2026-05", commit=True, refresh_import=True)

    assert result["status"] == "posted"
    assert result["legal_aid_debt_fee_total"] == 10000
    assert result["legal_aid_bonus_amount"] == 5000
    assert result["balance_after_laf_bonus"] == 15000
    assert result["case_bonus_pool"] == 7500
    assert result["case_bonus_employee_amount"] == 3750
    assert result["final_expense_total"] == 18750
    assert len(inserts) == 2
    assert "法扶消債酬金獎金" in inserts[0][5]
    assert "案件獎金" in inserts[1][5]


def test_monthly_bonus_route_preview(monkeypatch):
    from flask import Flask
    from flask_login import LoginManager, UserMixin

    from api.blueprints.osc_accounting import osc_accounting_bp

    app = Flask(__name__)
    app.config["TESTING"] = True
    app.config["LOGIN_DISABLED"] = True
    app.secret_key = "test"
    login = LoginManager(app)

    class User(UserMixin):
        id = "test"

    @login.user_loader
    def _load(_user_id):
        return User()

    app.register_blueprint(osc_accounting_bp)

    def fake_calc(**kwargs):
        assert kwargs["month"] == "2026-05"
        assert kwargs["commit"] is False
        return {"ok": True, "month": "2026-05", "status": "waiting_laf_fee"}

    monkeypatch.setattr("api.osc.accounting_bonus.calculate_monthly_bonus", fake_calc)
    resp = app.test_client().get("/api/osc/accounting/monthly-bonus?month=2026-05")
    assert resp.status_code == 200
    assert resp.get_json()["status"] == "waiting_laf_fee"


def test_accounting_transaction_allows_shared_expense_without_case(monkeypatch):
    from flask import Flask
    from flask_login import LoginManager, UserMixin

    from api.blueprints.osc_accounting import osc_accounting_bp

    app = Flask(__name__)
    app.config["TESTING"] = True
    app.config["LOGIN_DISABLED"] = True
    app.secret_key = "test"
    login = LoginManager(app)

    class User(UserMixin):
        id = "test"

    @login.user_loader
    def _load(_user_id):
        return User()

    app.register_blueprint(osc_accounting_bp)
    calls = []

    def fake_helpers():
        def fake_exec(sql, params=(), fetch="none"):
            calls.append((sql, params, fetch))
            return {"lastrowid": 123, "rowcount": 1}, {}

        return fake_exec, lambda x: x, lambda *a, **k: None, lambda x: None, lambda v, d=0: int(v or d)

    monkeypatch.setattr("api.blueprints.osc_accounting._get_osc_helpers", fake_helpers)
    resp = app.test_client().post(
        "/api/osc/accounting/transactions",
        json={
            "date": "2026-05-24",
            "type": "支出",
            "category": "薪資",
            "amount": 17000,
            "description": "法扶消債酬金獎金",
        },
    )
    assert resp.status_code == 200
    assert calls
    assert calls[0][1][0] is None


def test_accounting_transactions_xlsx_download(monkeypatch):
    from io import BytesIO

    from flask import Flask
    from flask_login import LoginManager, UserMixin
    from openpyxl import load_workbook

    from api.blueprints.osc_accounting import osc_accounting_bp

    app = Flask(__name__)
    app.config["TESTING"] = True
    app.config["LOGIN_DISABLED"] = True
    app.secret_key = "test"
    login = LoginManager(app)

    class User(UserMixin):
        id = "test"

    @login.user_loader
    def _load(_user_id):
        return User()

    app.register_blueprint(osc_accounting_bp)
    queries = []

    def fake_helpers():
        def fake_exec(sql, params=(), fetch="none"):
            queries.append((sql, params, fetch))
            if "FROM case_transactions t" in sql:
                assert "t.date >= %s" in sql
                assert "t.date <= %s" in sql
                return [
                    {
                        "id": 1,
                        "case_id": 10,
                        "case_number": "2026-0042",
                        "date": "2026-05-24",
                        "type": "收入",
                        "sub_type": "酬金",
                        "category": "法扶酬金",
                        "description": "林里法扶消債酬金",
                        "amount": 34000,
                    },
                    {
                        "id": 2,
                        "case_id": None,
                        "case_number": None,
                        "date": "2026-05-24",
                        "type": "支出",
                        "sub_type": "人事費",
                        "category": "薪資",
                        "description": "法扶消債酬金獎金",
                        "amount": 17000,
                    },
                ], {}
            return [], {}

        return fake_exec, lambda x: x, lambda *a, **k: None, lambda x: None, lambda v, d=0: int(v or d)

    monkeypatch.setattr("api.blueprints.osc_accounting._get_osc_helpers", fake_helpers)
    resp = app.test_client().get(
        "/api/osc/accounting/transactions/xlsx?start_date=2026-05-01&end_date=2026-05-31"
    )
    assert resp.status_code == 200
    assert resp.headers["Content-Disposition"].find("MAGI") >= 0
    assert resp.data.startswith(b"PK")

    wb = load_workbook(BytesIO(resp.data), data_only=True)
    assert wb.sheetnames == ["帳務明細", "摘要"]
    ws = wb["帳務明細"]
    assert ws["B2"].value == "2026-05-24"
    assert ws["C2"].value == "2026-0042"
    assert ws["H2"].value == "林里法扶消債酬金"
    summary = wb["摘要"]
    assert summary["B7"].value == 2
    assert summary["B8"].value == 34000
    assert summary["B9"].value == 17000
